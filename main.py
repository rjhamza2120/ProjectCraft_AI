import streamlit as st
import asyncio
from typing import Dict, List, Any
from dataclasses import dataclass
import json
import re
import io
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from simple_chat import simple_chat
from tools import ToolsMain
from theme import (
    add_custom_css, 
    create_animated_title, 
    create_progress_indicator, 
    create_interactive_assistant, 
    create_sidebar_stages
)
from langchain_groq import ChatGroq
from langchain.prompts import ChatPromptTemplate
from langchain_core.messages import HumanMessage, SystemMessage

import os
from dotenv import load_dotenv
load_dotenv()

@dataclass
class ProjectDetails:
    title: str
    short_description: str
    detailed_description: str
    components: List[Dict[str, str]]
    frameworks: List[str]
    youtube_links: List[str]
    github_repos: List[str]
    difficulty_level: str
    estimated_time: str

class ProjectGuideAssistant:
    def __init__(self):
        self.llm = ChatGroq(temperature=0.2, model="moonshotai/kimi-k2-instruct")
        self.simple_chat = simple_chat()
        self.tools = ToolsMain()
        self.tool_list = self.tools()
        self.tool_map = {t.name: t for t in self.tool_list}
        
        # Natural conversation prompt for project exploration
        self.refinement_prompt = ChatPromptTemplate.from_template("""
        You are a friendly project mentor who helps students and learners explore project ideas through natural conversation.
        
        User's input: {user_input}
        Previous conversation: {conversation_history}
        
        Your approach should be conversational and supportive:
        1. Have a natural, friendly chat about their interests and goals
        2. Ask open-ended questions to understand what they want to learn or build
        3. Suggest project ideas based on their interests and skill level
        4. Clarify doubts and provide guidance without being overly technical
        5. Only when they're excited about a specific project idea, encourage them to move forward
        
        Guidelines for conversation:
        - Be encouraging and enthusiastic about their ideas
        - Ask about their background, interests, and what they want to achieve
        - Suggest multiple project options and let them choose
        - Explain concepts in simple, understandable terms
        - Address any doubts or concerns they have
        - Make them feel confident about their project choice
        
        Examples of good responses:
        - "That sounds really interesting! What got you interested in this area?"
        - "Have you worked on anything similar before, or would this be your first project in this field?"
        - "I can think of a few different approaches. Would you prefer something more hands-on with hardware, or are you more interested in software development?"
        - "That's a great question! Let me explain how that would work..."
        
        Keep the conversation natural and engaging. Only suggest moving to the detailed planning phase when they're clearly excited about a specific project.
        """)
        
        # Trending projects generation prompt
        self.trending_projects_prompt = ChatPromptTemplate.from_template("""
        Generate 6 trending and popular project ideas for the field: {engineering_field}
        
        Focus on projects that are:
        1. Currently relevant and in-demand in the industry
        2. Educational and skill-building
        3. Suitable for different levels (beginner to advanced)
        4. Implementable with available resources
        5. Impressive for portfolios and resumes
        
        Provide a JSON response with this exact structure:
        {{
            "projects": [
                {{
                    "title": "Project Name",
                    "description": "Brief 2-3 sentence description of what the project does and why it's trending",
                    "difficulty": "Beginner/Intermediate/Advanced",
                    "category": "Semester Project/FYP/Hobby Project/Industry Project",
                    "key_technologies": ["Tech1", "Tech2", "Tech3"],
                    "why_trending": "Brief explanation of why this project is currently popular"
                }}
            ]
        }}
        
        Make sure projects are diverse in difficulty and application areas within the selected field.
        """)
        
        # Project-specific refinement prompt
        self.project_refinement_prompt = ChatPromptTemplate.from_template("""
        You are helping a student refine their project idea: {project_title}
        
        Current context:
        - Engineering Field: {engineering_field}
        - Project Type: {project_type}
        - Complexity Level: {complexity_level}
        - Previous responses: {user_responses}
        
        Ask ONE specific, focused question about this project to help refine it further.
        Focus on:
        1. Project scope and specific features they want to include
        2. Technical components, sensors, or libraries they prefer
        3. Target audience or use case scenarios
        4. Integration with other systems or platforms
        5. Performance requirements or constraints
        
        Keep questions practical and project-specific. Avoid asking about timelines, resources, or general preferences.
        Make the question conversational and helpful.
        
        Example good questions:
        - "What specific sensors would you like to use for monitoring? Temperature, humidity, soil moisture, or something else?"
        - "Do you want this to be a mobile app, web application, or desktop software?"
        - "Should this system work offline or do you want cloud integration for data storage?"
        - "What kind of user interface are you thinking - simple dashboard, mobile notifications, or voice control?"
        
        Respond with just the question, no additional text.
        """)
        
        # Project details generation prompt
        self.project_prompt = ChatPromptTemplate.from_template("""
        Generate a comprehensive project guide for: {project_title}
        
        Create a detailed, practical project plan that's educational and achievable.
        Focus on clear learning outcomes and step-by-step implementation.
        
        Provide a JSON response with this exact structure:
        {{
            "title": "Project Title",
            "short_description": "Clear, engaging one-sentence description of what the project does",
            "detailed_description": "Comprehensive guide including:\n1. Project overview and learning objectives\n2. Prerequisites and required knowledge\n3. Step-by-step implementation process\n4. Key concepts and technologies explained\n5. Testing and validation methods\n6. Potential extensions and improvements\n7. Common challenges and troubleshooting tips",
            "components": [
                {{"name": "Component Name", "purpose": "What this component does and why it's needed", "specs": "Detailed specifications and where to buy"}}
            ],
            "frameworks": ["Framework1", "Framework2", "Framework3"],
            "difficulty_level": "Beginner/Intermediate/Advanced",
            "estimated_time": "X weeks/months"
        }}
        
        Make it comprehensive yet approachable, with clear explanations of concepts and practical implementation steps.
        Include enough detail for someone to actually build the project successfully.
        """)

    async def generate_trending_projects(self, engineering_field: str) -> List[Dict]:
        """Generate trending projects for the selected engineering field"""
        try:
            response = self.llm.invoke(
                self.trending_projects_prompt.format_messages(engineering_field=engineering_field)
            )
            
            # Parse the response
            try:
                import json
                response_text = response.content if hasattr(response, 'content') else str(response)
                # Clean up the response to extract JSON
                if '```json' in response_text:
                    response_text = response_text.split('```json')[1].split('```')[0]
                elif '```' in response_text:
                    response_text = response_text.split('```')[1].split('```')[0]
                
                projects_data = json.loads(response_text.strip())
                return projects_data.get("projects", [])
            except:
                # Fallback trending projects
                return self._get_fallback_projects(engineering_field)
                
        except Exception as e:
            return self._get_fallback_projects(engineering_field)
    
    def _get_fallback_projects(self, engineering_field: str) -> List[Dict]:
        """Fallback trending projects if generation fails"""
        fallback_projects = {
            "💻 Computing & Software": [
                {"title": "AI Chatbot with RAG", "description": "Build an intelligent chatbot using Retrieval Augmented Generation for domain-specific queries.", "difficulty": "Intermediate", "category": "FYP", "key_technologies": ["Python", "LangChain", "OpenAI API"], "why_trending": "RAG systems are highly demanded in industry"},
                {"title": "Real-time Stock Price Predictor", "description": "Machine learning model that predicts stock prices using historical data and news sentiment.", "difficulty": "Advanced", "category": "FYP", "key_technologies": ["Python", "TensorFlow", "APIs"], "why_trending": "FinTech and ML integration is booming"},
                {"title": "Smart Home Automation App", "description": "Mobile app to control IoT devices with voice commands and automated schedules.", "difficulty": "Intermediate", "category": "Semester Project", "key_technologies": ["React Native", "IoT", "Firebase"], "why_trending": "Smart home market is rapidly growing"},
            ],
            "⚡ Electrical & Electronics": [
                {"title": "Solar Panel Monitoring System", "description": "IoT-based system to monitor solar panel efficiency and environmental conditions.", "difficulty": "Intermediate", "category": "FYP", "key_technologies": ["Arduino", "Sensors", "IoT"], "why_trending": "Renewable energy focus worldwide"},
                {"title": "Smart Grid Energy Management", "description": "System to optimize energy distribution and monitor power consumption in real-time.", "difficulty": "Advanced", "category": "FYP", "key_technologies": ["Microcontrollers", "Power Electronics", "Communication"], "why_trending": "Smart grid technology is essential for future"},
                {"title": "Wireless Phone Charging Pad", "description": "Design and build an efficient wireless charging system for mobile devices.", "difficulty": "Beginner", "category": "Semester Project", "key_technologies": ["Coils", "Power Electronics", "PCB"], "why_trending": "Wireless charging is standard in modern devices"},
            ]
        }
        return fallback_projects.get(engineering_field, fallback_projects["💻 Computing & Software"])
    
    async def ask_refinement_question(self, project_title: str, engineering_field: str, 
                                    project_type: str, complexity_level: str, user_responses: Dict) -> str:
        """Ask a specific refinement question about the selected project"""
        try:
            response = self.llm.invoke(
                self.project_refinement_prompt.format_messages(
                    project_title=project_title,
                    engineering_field=engineering_field,
                    project_type=project_type,
                    complexity_level=complexity_level,
                    user_responses=str(user_responses)
                )
            )
            
            question = response.content if hasattr(response, 'content') else str(response)
            return question.strip()
            
        except Exception as e:
            # Fallback questions based on project type
            fallback_questions = [
                f"What specific features would you like to include in your {project_title}?",
                f"Which technologies or components are you most interested in using for this {project_title}?",
                f"What would be the main use case or target audience for your {project_title}?",
                f"Do you want to focus more on the hardware side or software side of this {project_title}?",
                f"What makes your {project_title} unique compared to existing solutions?"
            ]
            import random
            return random.choice(fallback_questions)

    async def refine_project_idea(self, user_input: str, conversation_history: List[str]) -> str:
        """Refine the user's project idea through conversation"""
        try:
            # Use simple_chat for idea refinement with better context
            context = f"Previous conversation: {' '.join(conversation_history[-3:])}"
            refined_query = f"Help refine this project idea: {user_input}\nContext: {context}"
            
            response = await asyncio.create_task(
                asyncio.to_thread(self.simple_chat, refined_query)
            )
            return response
        except Exception as e:
            return f"I had trouble understanding that. Could you tell me more about what you'd like to build? For example, do you want to make something that helps around the house, or maybe something fun to play with?"

    async def generate_project_details(self, project_title: str) -> ProjectDetails:
        """Generate comprehensive project details with enhanced context"""
        try:
            # Get project context from session state
            engineering_field = getattr(st.session_state, 'selected_subdomain', None) or getattr(st.session_state, 'selected_field', '')
            project_type = getattr(st.session_state, 'project_type', 'General Project')
            complexity_level = getattr(st.session_state, 'complexity_level', 'Intermediate')
            user_responses = getattr(st.session_state, 'user_responses', {})
            
            # Enhanced prompt with context
            enhanced_prompt = f"""
            Generate a comprehensive project guide for: {project_title}
            
            Context:
            - Engineering Field: {engineering_field}
            - Project Type: {project_type}
            - Complexity Level: {complexity_level}
            - User Requirements: {'; '.join([f"{k}: {v}" for k, v in user_responses.items()])}
            
            Create a detailed, practical project plan that's educational and achievable.
            Focus on clear learning outcomes and step-by-step implementation.
            
            Provide a JSON response with this exact structure:
            {{
                "title": "{project_title}",
                "short_description": "Clear, engaging description of what the project does and its real-world applications",
                "detailed_description": "Comprehensive guide including:\\n1. Project overview and learning objectives\\n2. Prerequisites and required knowledge\\n3. Step-by-step implementation process with detailed explanations\\n4. Key concepts and technologies explained clearly\\n5. Testing and validation methods\\n6. Potential extensions and improvements\\n7. Common challenges and troubleshooting tips\\n8. Real-world applications and use cases",
                "components": [
                    {{"name": "Component Name", "purpose": "What this component does and why it's needed", "specs": "Detailed specifications, model numbers, and where to buy"}}
                ],
                "frameworks": ["Framework1", "Framework2", "Framework3"],
                "difficulty_level": "{complexity_level.split(' - ')[0] if ' - ' in complexity_level else complexity_level}",
                "estimated_time": "X weeks/months based on complexity"
            }}
            
            Make it comprehensive yet approachable, with clear explanations suitable for the specified complexity level.
            Include enough detail for someone to actually build the project successfully.
            """
            
            # Generate basic project structure
            project_response = self.llm.invoke([
                SystemMessage(content="You are an expert project mentor creating detailed, practical project guides."),
                HumanMessage(content=enhanced_prompt)
            ])
            
            # Parse the response with robust error handling
            project_data = self._parse_llm_response(project_response.content)
            
            if not project_data:
                # Create comprehensive fallback data
                project_data = {
                    "title": project_title,
                    "short_description": f"A comprehensive {engineering_field} project focusing on {project_title}",
                    "detailed_description": self._create_fallback_description(project_title, engineering_field, complexity_level),
                    "components": self._create_fallback_components(engineering_field),
                    "frameworks": self._create_fallback_frameworks(engineering_field),
                    "difficulty_level": complexity_level.split(' - ')[0] if ' - ' in complexity_level else complexity_level,
                    "estimated_time": "4-8 weeks"
                }
            
            # Use tools to get additional information with enhanced project context
            project_context = {
                'engineering_field': engineering_field,
                'user_responses': user_responses,
                'project_type': project_data.get('type', ''),
                'complexity_level': complexity_level
            }
            
            youtube_links = await self.get_youtube_tutorials(project_title, project_context)
            github_repos = await self.get_github_repos(project_title, engineering_field)
            
            return ProjectDetails(
                title=project_data.get("title", project_title),
                short_description=project_data.get("short_description", ""),
                detailed_description=project_data.get("detailed_description", ""),
                components=project_data.get("components", []),
                frameworks=project_data.get("frameworks", []),
                youtube_links=youtube_links,
                github_repos=github_repos,
                difficulty_level=project_data.get("difficulty_level", "Intermediate"),
                estimated_time=project_data.get("estimated_time", "4-6 weeks")
            )
            
        except Exception as e:
            st.error(f"Error generating project details: {str(e)}")
            return ProjectDetails(
                title=project_title,
                short_description=f"Custom {project_title} project",
                detailed_description=self._create_fallback_description(project_title, 
                    getattr(st.session_state, 'selected_field', 'Engineering'), 
                    getattr(st.session_state, 'complexity_level', 'Intermediate')),
                components=self._create_fallback_components(getattr(st.session_state, 'selected_field', 'Engineering')),
                frameworks=self._create_fallback_frameworks(getattr(st.session_state, 'selected_field', 'Engineering')),
                youtube_links=[],
                github_repos=[],
                difficulty_level="Intermediate",
                estimated_time="4-6 weeks"
            )
    
    def _create_fallback_description(self, project_title: str, engineering_field: str, complexity_level: str) -> str:
        """Create a comprehensive fallback description"""
        return f"""
## Project Overview
{project_title} is a {complexity_level.lower()} level project in {engineering_field} designed to provide hands-on learning experience and practical skills development.

## Learning Objectives
- Understand core concepts in {engineering_field}
- Develop practical implementation skills
- Learn industry-standard tools and practices
- Build a portfolio-worthy project

## Prerequisites
- Basic understanding of {engineering_field} principles
- Familiarity with relevant programming languages or tools
- Access to development environment and components

## Implementation Steps

### Phase 1: Planning and Setup
1. Review project requirements and scope
2. Set up development environment
3. Gather necessary components and tools
4. Create project structure and documentation

### Phase 2: Core Development
1. Implement basic functionality
2. Test individual components
3. Integrate system components
4. Optimize performance and reliability

### Phase 3: Testing and Refinement
1. Conduct comprehensive testing
2. Debug and fix issues
3. Enhance user experience
4. Document the solution

### Phase 4: Deployment and Documentation
1. Prepare for deployment
2. Create user documentation
3. Prepare presentation materials
4. Plan future enhancements

## Common Challenges and Solutions
- **Challenge**: Complex integration between components
  **Solution**: Break down into smaller, testable modules

- **Challenge**: Performance optimization
  **Solution**: Profile the system and identify bottlenecks

- **Challenge**: Debugging complex issues
  **Solution**: Use systematic debugging approaches and logging

## Real-World Applications
This project demonstrates skills and concepts directly applicable to:
- Industry projects in {engineering_field}
- Academic research and development
- Personal portfolio development
- Startup and entrepreneurial ventures

## Success Metrics
- Functional implementation meeting all requirements
- Clean, documented, and maintainable code
- Comprehensive testing and validation
- Professional presentation and documentation
        """
    
    def _create_fallback_components(self, engineering_field: str) -> List[Dict]:
        """Create fallback components based on engineering field"""
        component_map = {
            "💻 Computing & Software": [
                {"name": "Development Environment", "purpose": "IDE and development tools", "specs": "Visual Studio Code, Python 3.9+, Git"},
                {"name": "Libraries and Frameworks", "purpose": "Core software libraries", "specs": "Based on project requirements"},
                {"name": "Database System", "purpose": "Data storage and management", "specs": "SQLite, PostgreSQL, or MongoDB"}
            ],
            "⚡ Electrical & Electronics": [
                {"name": "Microcontroller", "purpose": "Central processing unit", "specs": "Arduino Uno, Raspberry Pi, or ESP32"},
                {"name": "Sensors", "purpose": "Data collection and monitoring", "specs": "Temperature, humidity, or motion sensors"},
                {"name": "Power Supply", "purpose": "System power management", "specs": "5V/3.3V regulated power supply"}
            ],
            "⚙️ Mechanical & Manufacturing": [
                {"name": "Mechanical Components", "purpose": "Physical system parts", "specs": "Gears, motors, or actuators"},
                {"name": "Control System", "purpose": "System automation", "specs": "PLC or microcontroller-based control"},
                {"name": "Measurement Tools", "purpose": "Precision measurement", "specs": "Calipers, gauges, or sensors"}
            ]
        }
        return component_map.get(engineering_field, component_map["💻 Computing & Software"])
    
    def _create_fallback_frameworks(self, engineering_field: str) -> List[str]:
        """Create fallback frameworks based on engineering field"""
        framework_map = {
            "💻 Computing & Software": ["Python", "JavaScript", "React", "Node.js", "SQL"],
            "⚡ Electrical & Electronics": ["Arduino IDE", "KiCad", "MATLAB", "LabVIEW", "Altium Designer"],
            "⚙️ Mechanical & Manufacturing": ["SolidWorks", "AutoCAD", "MATLAB", "LabVIEW", "Python"],
            "🏗️ Civil & Infrastructure": ["AutoCAD", "STAAD Pro", "ETABS", "Python", "MATLAB"],
            "⚗️ Chemical & Materials": ["MATLAB", "Aspen Plus", "ChemCAD", "Python", "R"]
        }
        return framework_map.get(engineering_field, ["Python", "MATLAB", "Documentation Tools"])

    def _parse_llm_response(self, response_content: str) -> dict:
        """Robust JSON parsing with multiple fallback attempts"""
        try:
            # Clean the response content
            content = response_content.strip()
            
            # Remove markdown code blocks
            if content.startswith('```json'):
                content = content[7:]
            elif content.startswith('```'):
                content = content[3:]
            
            if content.endswith('```'):
                content = content[:-3]
            
            # Try direct JSON parsing
            return json.loads(content)
            
        except json.JSONDecodeError:
            try:
                # Try to extract JSON from text using regex
                json_pattern = r'\{.*\}'
                matches = re.findall(json_pattern, content, re.DOTALL)
                if matches:
                    # Try the largest match (most likely to be complete)
                    largest_match = max(matches, key=len)
                    return json.loads(largest_match)
            except:
                pass
            
            try:
                # Try to fix common JSON issues
                # Remove trailing commas
                content = re.sub(r',(\s*[}\]])', r'\1', content)
                # Fix unescaped quotes in strings
                content = re.sub(r'(?<!\\)"(?![:,\]\}])', r'\\"', content)
                return json.loads(content)
            except:
                pass
        
        # If all parsing fails, return None to trigger fallback
        return None

    async def get_youtube_tutorials(self, project_title: str, project_context: Dict = None) -> List[str]:
        """Expert-level YouTube API integration with advanced filtering and project-specific search"""
        try:
            # Create enhanced tool instance with advanced YouTube API capabilities
            enhanced_youtube_tool = self._create_enhanced_youtube_tool()
            if not enhanced_youtube_tool:
                print("Enhanced YouTube tool creation failed")
                engineering_field = project_context.get('engineering_field', '') if project_context else ''
                return self._get_fallback_youtube_search_urls(project_title, engineering_field)
            
            # Extract comprehensive project context
            engineering_field = project_context.get('engineering_field', '') if project_context else ''
            user_responses = project_context.get('user_responses', {}) if project_context else {}
            project_type = project_context.get('project_type', '') if project_context else ''
            complexity_level = project_context.get('complexity_level', '') if project_context else ''
            
            # Generate expert-level search strategies
            search_strategies = self._generate_expert_search_strategies(
                project_title, engineering_field, user_responses, project_type, complexity_level
            )
            
            print(f"🔍 Using {len(search_strategies)} expert search strategies for: {project_title}")
            
            all_videos = []
            for strategy in search_strategies:
                try:
                    print(f"🚀 Executing search strategy: {strategy['name']}")
                    
                    # Execute advanced YouTube search with expert parameters
                    videos = await self._execute_advanced_youtube_search(
                        enhanced_youtube_tool, strategy, project_title, project_context
                    )
                    
                    if videos:
                        # Apply expert-level filtering and scoring
                        filtered_videos = self._apply_expert_video_filtering(
                            videos, project_title, project_context, strategy
                        )
                        all_videos.extend(filtered_videos)
                        print(f"✅ Strategy '{strategy['name']}' found {len(filtered_videos)} relevant videos")
                    
                except Exception as e:
                    print(f"❌ Error in search strategy '{strategy['name']}': {e}")
                    continue
            
            if all_videos:
                # Advanced video ranking and deduplication
                final_videos = self._rank_and_deduplicate_videos(all_videos, project_title, project_context)
                
                if final_videos:
                    print(f"🎯 Returning {len(final_videos)} highly relevant YouTube tutorials")
                    return [video['url'] for video in final_videos[:6]]  # Top 6 videos
            
            print("No valid YouTube videos found, returning enhanced fallback URLs")
                
        except Exception as e:
            print(f"💥 Critical error in YouTube tutorial search: {e}")
        
        # Enhanced fallback with project context
        engineering_field = project_context.get('engineering_field', '') if project_context else ''
        return self._get_enhanced_fallback_youtube_urls(project_title, project_context)
    
    def _create_enhanced_youtube_tool(self):
        """Create enhanced YouTube tool with advanced API capabilities"""
        try:
            fresh_tools = ToolsMain()
            fresh_tool_list = fresh_tools()
            fresh_tool_map = {t.name: t for t in fresh_tool_list}
            return fresh_tool_map.get("youtube_search")
        except Exception as e:
            print(f"Failed to create enhanced YouTube tool: {e}")
            return None
    
    def _generate_expert_search_strategies(self, project_title: str, engineering_field: str, 
                                         user_responses: Dict, project_type: str, complexity_level: str) -> List[Dict]:
        """Generate expert-level search strategies with advanced parameters and filtering"""
        strategies = []
        
        # Extract advanced project context
        tech_stack = self._extract_tech_stack(project_title, engineering_field, user_responses)
        domain_keywords = self._extract_domain_keywords(engineering_field, project_type)
        complexity_terms = self._get_complexity_terms(complexity_level)
        
        # Strategy 1: Deep Technical Tutorial Search
        strategies.append({
            'name': 'Deep Technical Tutorial',
            'query_template': f"{project_title} tutorial complete",
            'parameters': {
                'videoDuration': 'medium',
                'order': 'relevance'
            },
            'filters': {
                'min_duration': 240,  # 4+ minutes
                'exclude_shorts': True,
                'require_technical': True,
                'boost_code_tutorials': True
            }
        })
        
        # Strategy 2: Project-Specific Build Guide
        strategies.append({
            'name': 'Project Build Guide',
            'query_template': f"build {project_title} step by step",
            'parameters': {
                'videoDuration': 'long',
                'order': 'rating'
            },
            'filters': {
                'min_duration': 600,  # 10+ minutes
                'prefer_step_by_step': True,
                'boost_complete_projects': True,
                'require_implementation': True
            }
        })
        
        # Strategy 3: Domain-Specific Expert Content
        if engineering_field and domain_keywords:
            strategies.append({
                'name': 'Domain Expert Content',
                'query_template': f"{project_title} {engineering_field} guide",
                'parameters': {
                    'videoDuration': 'any',
                    'order': 'viewCount'
                },
                'filters': {
                    'boost_expert_channels': True,
                    'require_domain_match': True,
                    'min_views': 1000
                }
            })
        
        # Strategy 4: Latest Technology Implementation
        strategies.append({
            'name': 'Latest Tech Implementation',
            'query_template': f"{project_title} 2024 tutorial",
            'parameters': {
                'videoDuration': 'medium',
                'order': 'date'
            },
            'filters': {
                'prefer_recent': True,
                'boost_trending_tech': True,
                'require_updated_content': True
            }
        })
        
        return strategies
    
    def _extract_tech_stack(self, project_title: str, engineering_field: str, user_responses: Dict) -> List[str]:
        """Extract technology stack keywords from project context"""
        tech_stack = []
        
        # Common tech stack patterns
        tech_patterns = {
            'web': ['react', 'vue', 'angular', 'node', 'express', 'django', 'flask', 'javascript', 'python'],
            'mobile': ['react native', 'flutter', 'swift', 'kotlin', 'ionic', 'xamarin'],
            'data': ['python', 'pandas', 'numpy', 'tensorflow', 'pytorch', 'sql', 'mongodb'],
            'ai': ['machine learning', 'deep learning', 'tensorflow', 'pytorch', 'opencv', 'nlp'],
            'iot': ['arduino', 'raspberry pi', 'esp32', 'sensors', 'mqtt', 'bluetooth'],
            'automation': ['selenium', 'pytest', 'jenkins', 'docker', 'kubernetes'],
            'game': ['unity', 'unreal', 'c#', 'python', 'godot', 'blender']
        }
        
        project_lower = project_title.lower()
        field_lower = engineering_field.lower()
        
        # Match technology patterns
        for category, techs in tech_patterns.items():
            if category in project_lower or category in field_lower:
                tech_stack.extend(techs[:3])
        
        # Extract from user responses
        if user_responses:
            for response in user_responses.values():
                if isinstance(response, str):
                    response_lower = response.lower()
                    for category, techs in tech_patterns.items():
                        for tech in techs:
                            if tech in response_lower and tech not in tech_stack:
                                tech_stack.append(tech)
        
        return tech_stack[:8]  # Top 8 relevant technologies
    
    def _extract_domain_keywords(self, engineering_field: str, project_type: str) -> List[str]:
        """Extract domain-specific keywords for targeted search"""
        domain_map = {
            'computer': ['software', 'programming', 'development', 'coding', 'algorithm'],
            'electrical': ['electronics', 'circuit', 'embedded', 'microcontroller', 'arduino'],
            'mechanical': ['engineering', 'design', 'cad', 'modeling', 'simulation'],
            'civil': ['construction', 'structural', 'design', 'autocad', 'engineering'],
            'aerospace': ['aerodynamics', 'flight', 'aircraft', 'aerospace', 'simulation'],
            'chemical': ['process', 'chemical', 'reaction', 'plant', 'engineering'],
            'biomedical': ['medical', 'biomedical', 'healthcare', 'device', 'signal'],
            'robotics': ['robot', 'robotics', 'automation', 'control', 'sensors']
        }
        
        field_lower = engineering_field.lower()
        keywords = []
        
        for domain, terms in domain_map.items():
            if domain in field_lower:
                keywords.extend(terms[:3])
        
        # Add project type specific terms
        if project_type:
            type_terms = {
                'web app': ['web', 'application', 'frontend', 'backend'],
                'mobile app': ['mobile', 'app', 'android', 'ios'],
                'desktop app': ['desktop', 'application', 'gui', 'interface'],
                'api': ['api', 'rest', 'endpoint', 'service'],
                'database': ['database', 'sql', 'data', 'storage'],
                'game': ['game', 'gaming', 'graphics', 'engine']
            }
            
            for proj_type, terms in type_terms.items():
                if proj_type in project_type.lower():
                    keywords.extend(terms[:2])
        
        return list(set(keywords))[:6]  # Top 6 unique domain keywords
    
    def _get_complexity_terms(self, complexity_level: str) -> List[str]:
        """Get complexity-appropriate search terms"""
        complexity_map = {
            'beginner': ['beginner', 'basic', 'introduction', 'getting started', 'simple'],
            'intermediate': ['intermediate', 'practical', 'hands-on', 'project-based'],
            'advanced': ['advanced', 'expert', 'professional', 'production', 'enterprise'],
            'expert': ['expert', 'advanced', 'master', 'professional', 'industry']
        }
        
        if complexity_level and complexity_level.lower() in complexity_map:
            return complexity_map[complexity_level.lower()][:3]
        
        return ['tutorial', 'guide', 'course']  # Default terms
    
    async def _execute_advanced_youtube_search(self, youtube_tool, strategy: Dict, 
                                             project_title: str, project_context: Dict) -> List[Dict]:
        """Execute advanced YouTube search with expert parameters"""
        try:
            # Build advanced query with strategy parameters
            query = strategy['query_template']
            parameters = strategy.get('parameters', {})
            
            print(f"🔍 Advanced search query: {query}")
            print(f"📊 Search parameters: {parameters}")
            
            # Execute search with enhanced error handling
            result = await asyncio.to_thread(youtube_tool.invoke, {"query": query})
            
            if isinstance(result, str) and not result.startswith("Error") and "youtube.com" in result:
                # Parse and structure video data
                videos = self._parse_advanced_youtube_response(result, strategy, project_title)
                return videos
            else:
                print(f"⚠️ No valid results from YouTube API: {result}")
                return []
                
        except Exception as e:
            print(f"❌ Advanced YouTube search failed: {e}")
            return []
    
    def _parse_advanced_youtube_response(self, response: str, strategy: Dict, project_title: str) -> List[Dict]:
        """Parse YouTube API response with advanced video data extraction"""
        videos = []
        
        try:
            # Extract video URLs with enhanced patterns
            youtube_urls = self._extract_youtube_urls(response)
            
            # Parse video information from response text
            lines = response.split('\n')
            current_video = {}
            
            for line in lines:
                line = line.strip()
                
                # Extract title information
                if 'Title:' in line or 'title:' in line.lower():
                    title = line.split(':', 1)[1].strip() if ':' in line else line
                    current_video['title'] = title
                
                # Extract duration information
                elif 'Duration:' in line or 'duration:' in line.lower():
                    duration = line.split(':', 1)[1].strip() if ':' in line else ''
                    current_video['duration'] = self._parse_duration(duration)
                
                # Extract view count
                elif 'Views:' in line or 'views:' in line.lower():
                    views = line.split(':', 1)[1].strip() if ':' in line else ''
                    current_video['views'] = self._parse_view_count(views)
                
                # Extract channel information
                elif 'Channel:' in line or 'channel:' in line.lower():
                    channel = line.split(':', 1)[1].strip() if ':' in line else ''
                    current_video['channel'] = channel
                
                # Extract description
                elif 'Description:' in line or 'description:' in line.lower():
                    description = line.split(':', 1)[1].strip() if ':' in line else ''
                    current_video['description'] = description
                
                # When we find a URL, complete the video entry
                elif any(url in line for url in youtube_urls):
                    for url in youtube_urls:
                        if url in line:
                            current_video['url'] = url
                            current_video['strategy'] = strategy['name']
                            
                            # Calculate advanced relevance score
                            current_video['relevance_score'] = self._calculate_advanced_relevance_score(
                                current_video, project_title, strategy
                            )
                            
                            videos.append(current_video.copy())
                            current_video = {}
                            break
            
            # Handle any remaining URLs without full metadata
            for url in youtube_urls:
                if not any(video['url'] == url for video in videos):
                    videos.append({
                        'url': url,
                        'title': 'YouTube Video',
                        'strategy': strategy['name'],
                        'relevance_score': 50  # Default score
                    })
            
        except Exception as e:
            print(f"Error parsing advanced YouTube response: {e}")
            # Fallback to basic URL extraction
            youtube_urls = self._extract_youtube_urls(response)
            for url in youtube_urls:
                videos.append({
                    'url': url,
                    'title': 'YouTube Tutorial',
                    'strategy': strategy['name'],
                    'relevance_score': 40
                })
        
        return videos
    
    def _parse_duration(self, duration_str: str) -> int:
        """Parse video duration string to seconds"""
        try:
            if not duration_str:
                return 0
            
            # Handle various duration formats
            duration_str = duration_str.lower().replace('duration:', '').strip()
            
            if 'min' in duration_str:
                parts = duration_str.split()
                minutes = int(parts[0]) if parts and parts[0].isdigit() else 0
                return minutes * 60
            elif ':' in duration_str:
                parts = duration_str.split(':')
                if len(parts) == 2:
                    minutes, seconds = parts
                    return int(minutes) * 60 + int(seconds)
                elif len(parts) == 3:
                    hours, minutes, seconds = parts
                    return int(hours) * 3600 + int(minutes) * 60 + int(seconds)
            
            return 0
        except:
            return 0
    
    def _parse_view_count(self, views_str: str) -> int:
        """Parse view count string to integer"""
        try:
            if not views_str:
                return 0
            
            views_str = views_str.lower().replace('views:', '').replace(',', '').strip()
            
            if 'k' in views_str:
                return int(float(views_str.replace('k', '')) * 1000)
            elif 'm' in views_str:
                return int(float(views_str.replace('m', '')) * 1000000)
            elif views_str.isdigit():
                return int(views_str)
            
            return 0
        except:
            return 0
    
    def _calculate_advanced_relevance_score(self, video: Dict, project_title: str, strategy: Dict) -> int:
        """Calculate advanced relevance score with multiple factors"""
        score = 0
        title = video.get('title', '').lower()
        description = video.get('description', '').lower()
        duration = video.get('duration', 0)
        views = video.get('views', 0)
        
        # Title relevance (40% weight)
        project_words = project_title.lower().split()
        title_matches = sum(10 for word in project_words if word in title)
        score += title_matches
        
        # Tutorial quality indicators (30% weight)
        quality_terms = ['tutorial', 'complete', 'full course', 'step by step', 'guide', 'build', 'create']
        quality_score = sum(8 for term in quality_terms if term in title)
        score += quality_score
        
        # Technical depth indicators (20% weight)
        tech_terms = ['implementation', 'code', 'programming', 'development', 'project']
        tech_score = sum(6 for term in tech_terms if term in title or term in description)
        score += tech_score
        
        # Duration bonus (5% weight)
        if 300 <= duration <= 3600:  # 5-60 minutes ideal
            score += 10
        elif duration > 3600:  # Long form content
            score += 15
        
        # View count factor (5% weight)
        if views > 10000:
            score += 8
        elif views > 1000:
            score += 5
        
        # Strategy-specific bonuses
        strategy_filters = strategy.get('filters', {})
        if strategy_filters.get('boost_complete_projects') and 'complete' in title:
            score += 15
        if strategy_filters.get('boost_code_tutorials') and ('code' in title or 'programming' in title):
            score += 12
        
        return score
    
    def _apply_expert_video_filtering(self, videos: List[Dict], project_title: str, 
                                    project_context: Dict, strategy: Dict) -> List[Dict]:
        """Apply expert-level filtering with advanced criteria"""
        filtered_videos = []
        strategy_filters = strategy.get('filters', {})
        
        for video in videos:
            # Apply duration filters
            duration = video.get('duration', 0)
            min_duration = strategy_filters.get('min_duration', 0)
            
            if min_duration > 0 and duration < min_duration:
                continue
            
            # Exclude YouTube Shorts
            if strategy_filters.get('exclude_shorts', False):
                if duration < 60 or '#shorts' in video.get('title', '').lower():
                    continue
            
            # Technical content requirement
            if strategy_filters.get('require_technical', False):
                title = video.get('title', '').lower()
                description = video.get('description', '').lower()
                tech_terms = ['code', 'programming', 'implementation', 'tutorial', 'guide', 'build']
                
                if not any(term in title or term in description for term in tech_terms):
                    continue
            
            # Domain matching requirement
            if strategy_filters.get('require_domain_match', False):
                engineering_field = project_context.get('engineering_field', '') if project_context else ''
                if engineering_field:
                    domain_keywords = self._extract_domain_keywords(engineering_field, '')
                    title_desc = (video.get('title', '') + ' ' + video.get('description', '')).lower()
                    
                    if not any(keyword in title_desc for keyword in domain_keywords):
                        continue
            
            # Minimum views requirement
            min_views = strategy_filters.get('min_views', 0)
            if min_views > 0 and video.get('views', 0) < min_views:
                continue
            
            # Apply relevance score threshold
            if video.get('relevance_score', 0) < 30:  # Minimum relevance threshold
                continue
            
            filtered_videos.append(video)
        
        return filtered_videos
    
    def _rank_and_deduplicate_videos(self, videos: List[Dict], project_title: str, 
                                   project_context: Dict) -> List[Dict]:
        """Advanced ranking and deduplication with multiple criteria"""
        if not videos:
            return []
        
        # Remove duplicates by URL
        seen_urls = set()
        unique_videos = []
        
        for video in videos:
            url = video.get('url', '')
            if url and url not in seen_urls:
                unique_videos.append(video)
                seen_urls.add(url)
        
        # Advanced ranking with multiple factors
        def calculate_final_score(video):
            base_score = video.get('relevance_score', 0)
            
            # Boost high-quality channels
            channel = video.get('channel', '').lower()
            quality_channels = ['traversy media', 'code with mosh', 'programming with mosh', 
                              'freecodecamp', 'the net ninja', 'academind', 'clever programmer']
            
            if any(ch in channel for ch in quality_channels):
                base_score += 20
            
            # Boost comprehensive tutorials
            title = video.get('title', '').lower()
            if any(term in title for term in ['complete', 'full course', 'comprehensive', 'masterclass']):
                base_score += 15
            
            # Boost recent content
            if video.get('strategy') == 'Latest Tech Implementation':
                base_score += 10
            
            # Duration sweet spot bonus
            duration = video.get('duration', 0)
            if 600 <= duration <= 3600:  # 10-60 minutes
                base_score += 12
            
            return base_score
        
        # Sort by final score
        unique_videos.sort(key=calculate_final_score, reverse=True)
        
        return unique_videos[:8]  # Return top 8 videos
    
    def _get_enhanced_fallback_youtube_urls(self, project_title: str, project_context: Dict) -> List[str]:
        """Enhanced fallback URLs with project context - return direct video links"""
        
        # Try to get direct video links using a simpler approach
        try:
            youtube_tool = self._create_enhanced_youtube_tool()
            if youtube_tool:
                # Simple search for direct links
                simple_query = f"{project_title} tutorial"
                result = youtube_tool.invoke({"query": simple_query})
                
                # Extract direct video URLs from the result
                if isinstance(result, str) and "youtube.com/watch" in result:
                    urls = self._extract_youtube_urls(result)
                    if urls:
                        return urls[:4]  # Return up to 4 direct video links
                        
        except Exception as e:
            print(f"Failed to get direct video links: {e}")
        
        # Fallback to search URLs only if direct links fail
        engineering_field = project_context.get('engineering_field', '') if project_context else ''
        base_queries = [
            f"{project_title} tutorial",
            f"{project_title} complete guide", 
            f"how to build {project_title}",
            f"{project_title} {engineering_field} tutorial"
        ]
        
        fallback_urls = []
        for query in base_queries:
            encoded_query = query.replace(' ', '+')
            fallback_urls.append(f"https://www.youtube.com/results?search_query={encoded_query}")
        
        return fallback_urls[:4]

    def _is_quality_tutorial_video(self, url: str, title: str, project_title: str) -> bool:
        """Enhanced filtering for quality tutorial videos"""
        if not url or not url.startswith('http'):
            return False
            
        title_lower = title.lower()
        
        # Exclude shorts and low-quality content
        exclusion_terms = ['shorts', '#shorts', 'meme', 'funny', 'react', 'tiktok', 'quick tip']
        if any(term in title_lower for term in exclusion_terms):
            return False
            
        # Prefer tutorial content
        tutorial_indicators = ['tutorial', 'how to', 'guide', 'build', 'create', 'make', 'step by step', 
                             'complete', 'full', 'beginner', 'learn', 'course', 'project']
        tutorial_score = sum(1 for indicator in tutorial_indicators if indicator in title_lower)
        
        # Check project relevance
        project_words = project_title.lower().split()
        relevance_score = sum(1 for word in project_words if word in title_lower)
        
        return tutorial_score >= 1 or relevance_score >= len(project_words) * 0.6
    
    def _calculate_video_relevance_score(self, title: str, project_title: str, engineering_field: str) -> int:
        """Calculate relevance score for video ranking"""
        score = 0
        title_lower = title.lower()
        
        # High-value tutorial indicators
        premium_terms = ['complete tutorial', 'full course', 'step by step', 'comprehensive guide', 
                        'from scratch', 'beginner to advanced']
        for term in premium_terms:
            if term in title_lower:
                score += 15
                
        # Standard tutorial terms
        tutorial_terms = ['tutorial', 'how to', 'guide', 'build', 'create', 'learn']
        score += sum(5 for term in tutorial_terms if term in title_lower)
        
        # Project title relevance
        project_words = project_title.lower().split()
        score += sum(10 for word in project_words if word in title_lower)
        
        # Engineering field relevance
        if engineering_field and any(word in title_lower for word in engineering_field.lower().split()):
            score += 8
            
        # Quality indicators
        quality_terms = ['professional', 'expert', 'advanced', 'industry', 'best practices']
        score += sum(3 for term in quality_terms if term in title_lower)
        
        return score
    
    def _extract_project_keywords(self, project_title: str, engineering_field: str) -> List[str]:
        """Extract relevant keywords from project title and field for better searches"""
        keywords = []
        
        # Clean and split project title
        title_words = [word.strip() for word in project_title.lower().split() if len(word) > 2]
        keywords.extend(title_words)
        
        # Add highly specific technical keywords based on field and project context
        field_tech_map = {
            'computer': ['programming', 'software', 'development', 'coding', 'algorithm', 'database', 'api', 'frontend', 'backend'],
            'electrical': ['circuit', 'electronics', 'arduino', 'microcontroller', 'sensors', 'embedded', 'iot', 'pcb'],
            'mechanical': ['design', 'modeling', 'simulation', 'cad', 'solidworks', 'autocad', 'engineering'],
            'civil': ['construction', 'structural', 'analysis', 'design', 'autocad', 'concrete', 'steel'],
            'aerospace': ['aerodynamics', 'flight', 'aircraft', 'simulation', 'cfd', 'matlab'],
            'chemical': ['process', 'reactor', 'simulation', 'aspen', 'plant', 'distillation'],
            'biomedical': ['medical', 'device', 'healthcare', 'signal', 'processing', 'matlab'],
            'environmental': ['sustainability', 'renewable', 'solar', 'water', 'treatment'],
            'robotics': ['robot', 'automation', 'control', 'sensors', 'actuators', 'ros'],
            'artificial intelligence': ['ai', 'machine learning', 'neural network', 'deep learning', 'python', 'tensorflow'],
            'data science': ['data', 'analytics', 'visualization', 'python', 'pandas', 'numpy', 'matplotlib']
        }
        
        # More precise field matching
        field_lower = engineering_field.lower()
        for field_key, tech_terms in field_tech_map.items():
            if field_key in field_lower or any(word in field_key for word in field_lower.split()):
                keywords.extend(tech_terms[:4])  # Take top 4 most relevant terms
                break
        
        # Add context-specific keywords based on project title patterns
        title_lower = project_title.lower()
        if 'web' in title_lower or 'website' in title_lower:
            keywords.extend(['html', 'css', 'javascript', 'react', 'nodejs'])
        elif 'mobile' in title_lower or 'app' in title_lower:
            keywords.extend(['android', 'ios', 'flutter', 'react native'])
        elif 'iot' in title_lower or 'smart' in title_lower:
            keywords.extend(['arduino', 'raspberry pi', 'sensors', 'wifi', 'bluetooth'])
        elif 'machine learning' in title_lower or 'ai' in title_lower:
            keywords.extend(['python', 'scikit-learn', 'tensorflow', 'keras'])
        elif 'database' in title_lower or 'management' in title_lower:
            keywords.extend(['sql', 'mysql', 'postgresql', 'mongodb'])
        
        return list(set(keywords))[:10]  # Return unique keywords, max 10
    
    def _extract_youtube_urls(self, result_text: str) -> List[str]:
        """Extract YouTube URLs using multiple patterns"""
        urls = []
        
        # Pattern 1: Standard YouTube watch URLs
        pattern1 = r'https://www\.youtube\.com/watch\?v=[a-zA-Z0-9_-]{11}'
        urls.extend(re.findall(pattern1, result_text))
        
        # Pattern 2: YouTube URLs with additional parameters
        pattern2 = r'https://www\.youtube\.com/watch\?v=[a-zA-Z0-9_-]+[^\s]*'
        additional_urls = re.findall(pattern2, result_text)
        for url in additional_urls:
            # Clean URL by removing extra characters
            clean_url = url.split('&')[0] if '&' in url else url
            if clean_url not in urls:
                urls.append(clean_url)
        
        # Pattern 3: Extract from formatted lines
        lines = result_text.split('\n')
        for line in lines:
            if 'youtube.com/watch' in line:
                # Extract URL from the line
                url_match = re.search(r'https://www\.youtube\.com/watch\?v=[a-zA-Z0-9_-]+', line)
                if url_match and url_match.group() not in urls:
                    urls.append(url_match.group())
        
        return list(set(urls))  # Remove duplicates
    
    def _parse_youtube_response(self, result_text: str, urls: List[str], project_title: str, engineering_field: str) -> List[Dict]:
        """Parse YouTube tool response to extract video details"""
        videos = []
        lines = result_text.split('\n')
        
        for url in urls:
            title = "Tutorial Video"
            description = ""
            
            # Find the line containing this URL
            for i, line in enumerate(lines):
                if url in line:
                    # Look backwards for title (usually in format "- Title by Channel")
                    for j in range(max(0, i-3), i):
                        if lines[j].startswith('- ') and 'by' in lines[j]:
                            title = lines[j][2:].split(' by ')[0].strip()
                            break
                    
                    # Look forward for description
                    if i + 1 < len(lines) and not lines[i + 1].startswith(('- ', 'http')):
                        description = lines[i + 1].strip()
                    
                    break
            
            # Quality filtering
            if self._is_quality_tutorial_video(url, title, project_title):
                score = self._calculate_video_relevance_score(title, project_title, engineering_field)
                videos.append({
                    'url': url,
                    'title': title,
                    'description': description,
                    'score': score
                })
        
        return videos
    
    def _get_fallback_youtube_search_urls(self, project_title: str, engineering_field: str) -> List[str]:
        """Generate fallback YouTube search URLs"""
        base_url = "https://www.youtube.com/results?search_query="
        search_terms = [
            f"{project_title.replace(' ', '+')}+tutorial+complete",
            f"{project_title.replace(' ', '+')}+step+by+step+guide",
            f"{engineering_field.replace(' ', '+')}+{project_title.replace(' ', '+')}+project",
            f"how+to+build+{project_title.replace(' ', '+')}+from+scratch",
            f"{project_title.replace(' ', '+')}+programming+tutorial+beginner"
        ]
        return [base_url + term for term in search_terms[:5]]

    async def get_github_repos(self, project_title: str, engineering_field: str = "") -> List[str]:
        """Get relevant GitHub repository links with enhanced project-specific filtering"""
        try:
            # Create fresh tool instance to avoid any potential caching
            fresh_tools = ToolsMain()
            fresh_tool_list = fresh_tools()
            fresh_tool_map = {t.name: t for t in fresh_tool_list}
            github_tool = fresh_tool_map.get("github_search")
            if not github_tool:
                print("GitHub tool not found in tool_map")
                return self._get_fallback_github_search_urls(project_title, engineering_field)
            
            # Create highly specific search queries based on project context
            project_keywords = self._extract_project_keywords(project_title, engineering_field)
            
            # Build very targeted search queries with technical specificity
            primary_keywords = project_keywords[:3] if len(project_keywords) >= 3 else project_keywords
            secondary_keywords = project_keywords[3:6] if len(project_keywords) > 3 else project_keywords
            
            search_queries = [
                f"{project_title} {engineering_field} source code implementation {' '.join(primary_keywords)}",
                f"{project_title} complete project repository {' '.join(secondary_keywords)} github",
                f"{' '.join(primary_keywords[:2])} {project_title} example implementation source",
                f"{project_title} {engineering_field} {' '.join(project_keywords[:2])} complete project code"
            ]
            
            # Add variety to search queries to avoid cached results
            enhanced_queries = []
            for i, query in enumerate(search_queries[:2]):
                if i == 0:
                    enhanced_queries.append(f"{query} open source project")
                else:
                    enhanced_queries.append(f"{query} complete source code")
            
            print(f"Searching GitHub with enhanced queries: {enhanced_queries}")
            
            all_repos = []
            for query in enhanced_queries:  # Use enhanced queries
                try:
                    print(f"Calling GitHub tool with query: {query}")
                    result = await asyncio.to_thread(github_tool.invoke, {"query": query})
                    print(f"GitHub tool result type: {type(result)}")
                    print(f"GitHub tool result length: {len(str(result)) if result else 0}")
                    print(f"GitHub tool result preview: {str(result)[:300]}...")
                    print(f"GitHub API call completed for project: {project_title}")
                    
                    if isinstance(result, str) and not result.startswith("Error") and "github.com" in result:
                        # Enhanced URL extraction and parsing
                        github_repos = self._parse_github_response(result, project_title, engineering_field)
                        all_repos.extend(github_repos)
                        print(f"Extracted {len(github_repos)} GitHub repositories")
                        
                    else:
                        print(f"GitHub tool returned error or no results: {result}")
                        
                except Exception as e:
                    print(f"Error processing GitHub query '{query}': {e}")
                    continue
            
            if all_repos:
                # Sort by relevance score and return top unique repositories
                all_repos.sort(key=lambda x: x.get('score', 0), reverse=True)
                unique_repos = []
                seen_urls = set()
                
                for repo in all_repos:
                    url = repo.get('url', '')
                    clean_url = self._clean_github_url(url)
                    if clean_url and clean_url not in seen_urls and len(unique_repos) < 5:
                        unique_repos.append(clean_url)
                        seen_urls.add(clean_url)
                        print(f"Added repo: {repo.get('name', 'Unknown')} - {clean_url}")
                
                if unique_repos:
                    return unique_repos
            
            print("No valid GitHub repositories found, returning fallback URLs")
                
        except Exception as e:
            print(f"Error fetching GitHub repositories: {e}")
        
        # Return fallback search URLs
        return self._get_fallback_github_search_urls(project_title, engineering_field)
    
    def _parse_github_response(self, result_text: str, project_title: str, engineering_field: str) -> List[Dict]:
        """Parse GitHub tool response to extract repository details"""
        repos = []
        
        # Extract GitHub URLs first
        github_urls = self._extract_github_urls(result_text)
        
        # Split into sections (repositories are usually separated by double newlines)
        sections = result_text.split('\n\n')
        
        for section in sections:
            if 'github.com' not in section:
                continue
                
            lines = [line.strip() for line in section.split('\n') if line.strip()]
            
            repo_url = None
            repo_name = "Repository"
            description = ""
            stars = 0
            
            for line in lines:
                # Extract repository URL
                if 'github.com' in line and line.startswith('http'):
                    repo_url = line.strip()
                # Extract repository name and stars (format: "- owner/repo ⭐ (123 stars)")
                elif line.startswith('- ') and '⭐' in line:
                    parts = line[2:].split(' ⭐')
                    if parts:
                        repo_name = parts[0].strip()
                        # Extract star count if available
                        if len(parts) > 1 and 'stars' in parts[1]:
                            star_match = re.search(r'\((\d+)\s*stars?\)', parts[1])
                            if star_match:
                                stars = int(star_match.group(1))
                # Extract description (usually the longest non-URL, non-title line)
                elif (not line.startswith(('http', '- ')) and 
                      len(line) > 20 and 
                      not line.isdigit() and
                      '⭐' not in line):
                    if len(line) > len(description):  # Take the longest description
                        description = line
            
            # Validate and score the repository
            if repo_url and self._is_quality_repository(repo_url, description, project_title):
                score = self._calculate_repo_relevance_score(repo_url, description, project_title, engineering_field)
                # Boost score based on stars
                if stars > 100:
                    score += 5
                elif stars > 50:
                    score += 3
                elif stars > 10:
                    score += 1
                
                repos.append({
                    'url': repo_url,
                    'name': repo_name,
                    'description': description,
                    'stars': stars,
                    'score': score
                })
        
        return repos
    
    def _extract_github_urls(self, result_text: str) -> List[str]:
        """Extract GitHub repository URLs using multiple patterns"""
        urls = []
        
        # Pattern 1: Standard GitHub repository URLs
        pattern1 = r'https://github\.com/[a-zA-Z0-9_.-]+/[a-zA-Z0-9_.-]+'
        urls.extend(re.findall(pattern1, result_text))
        
        # Pattern 2: Extract from formatted lines
        lines = result_text.split('\n')
        for line in lines:
            if 'github.com' in line and line.strip().startswith('http'):
                # Clean URL by removing trailing punctuation
                url = line.strip().rstrip('.,;')
                if url not in urls:
                    urls.append(url)
        
        return list(set(urls))  # Remove duplicates

    def _is_quality_repository(self, url: str, description: str, project_title: str) -> bool:
        """Check if repository is relevant and high-quality"""
        if not url or not url.startswith('http'):
            return False
            
        # Must be a main repository URL (not file, issue, or wiki)
        if any(path in url for path in ['/blob/', '/issues/', '/wiki/', '/releases/', '/pull/']):
            return False
            
        # Check URL structure (should be github.com/user/repo)
        url_parts = url.replace('https://github.com/', '').split('/')
        if len(url_parts) < 2:
            return False
            
        description_lower = description.lower()
        project_words = project_title.lower().split()
        
        # Check project relevance
        relevance_score = sum(1 for word in project_words if word in description_lower or word in url.lower())
        
        # Avoid low-quality repositories
        low_quality_indicators = ['test', 'hello-world', 'practice', 'homework', 'assignment', 'copy', 'clone']
        if any(indicator in description_lower for indicator in low_quality_indicators):
            return relevance_score >= len(project_words) * 0.8  # Higher threshold for potentially low-quality repos
            
        return relevance_score >= len(project_words) * 0.4
    
    def _calculate_repo_relevance_score(self, url: str, description: str, project_title: str, engineering_field: str) -> int:
        """Calculate relevance score for repository ranking"""
        score = 0
        description_lower = description.lower()
        url_lower = url.lower()
        
        # Project title relevance (high weight)
        project_words = project_title.lower().split()
        exact_matches = sum(5 for word in project_words if word in description_lower)
        partial_matches = sum(3 for word in project_words if word in url_lower)
        score += exact_matches + partial_matches
        
        # Engineering field relevance
        if engineering_field:
            field_words = engineering_field.lower().split()
            score += sum(3 for word in field_words if word in description_lower)
            
        # Quality indicators
        quality_indicators = ['complete', 'full', 'implementation', 'tutorial', 'example', 'demo', 
                            'production', 'professional', 'advanced', 'comprehensive']
        score += sum(4 for indicator in quality_indicators if indicator in description_lower)
        
        # Repository activity indicators (based on description patterns)
        activity_indicators = ['updated', 'maintained', 'active', 'popular', 'stars', 'contributors']
        score += sum(2 for indicator in activity_indicators if indicator in description_lower)
        
        # Prefer original implementations over forks (if detectable)
        if 'fork' in description_lower or 'forked' in description_lower:
            score -= 5
            
        # Bonus for detailed descriptions
        if len(description) > 50:
            score += 3
            
        return max(0, score)  # Ensure non-negative score
    
    def _clean_github_url(self, url: str) -> str:
        """Clean GitHub URL to get main repository URL"""
        if not url.startswith('http'):
            return url
            
        # Remove any trailing paths to get clean repo URL
        parts = url.replace('https://github.com/', '').split('/')
        if len(parts) >= 2:
            return f"https://github.com/{parts[0]}/{parts[1]}"
        return url
    
    def _get_fallback_github_search_urls(self, project_title: str, engineering_field: str) -> List[str]:
        """Generate fallback GitHub URLs - try to get direct repo links first"""
        
        # Try to get direct repository links using the GitHub tool
        try:
            fresh_tools = ToolsMain()
            fresh_tool_list = fresh_tools()
            fresh_tool_map = {t.name: t for t in fresh_tool_list}
            github_tool = fresh_tool_map.get("github_search")
            
            if github_tool:
                # Simple search for direct repo links
                simple_query = f"{project_title} {engineering_field} project"
                result = github_tool.invoke({"query": simple_query})
                
                # Extract direct repository URLs from the result
                if isinstance(result, str) and "github.com/" in result:
                    urls = self._extract_github_urls(result)
                    if urls:
                        print(f"✅ Found {len(urls)} direct GitHub repository links")
                        return urls[:6]  # Return up to 6 direct repo links
                        
        except Exception as e:
            print(f"⚠️ Failed to get direct repo links: {e}")
        
        # Fallback to search URLs only if direct links fail
        print("⚠️ Using GitHub search URLs as fallback")
        base_url = "https://github.com/search?q="
        search_terms = [
            f"{project_title.replace(' ', '+')}+project+implementation",
            f"{project_title.replace(' ', '+')}+source+code+example", 
            f"{engineering_field.replace(' ', '+')}+{project_title.replace(' ', '+')}",
            f"{project_title.replace(' ', '+')}+tutorial+repository",
            f"awesome+{project_title.replace(' ', '+')}+projects",
            f"{project_title.replace(' ', '+')}+complete+implementation"
        ]
        return [base_url + term + "&type=repositories" for term in search_terms]

    def generate_excel_guide(self, project_details: ProjectDetails, user_name: str = "Builder") -> bytes:
        """Generate a professional Excel project guide with enhanced formatting and tables"""
        try:
            # Create a new workbook and get active worksheet
            wb = Workbook()
            
            # Remove default sheet and create named sheets
            wb.remove(wb.active)
            
            # Define professional color scheme
            primary_color = "1E40AF"  # Blue
            secondary_color = "F8FAFC"  # Light gray
            accent_color = "10B981"  # Green
            text_color = "374151"  # Dark gray
            
            # Define common styles
            header_font = Font(name="Calibri", size=16, bold=True, color=primary_color)
            subheader_font = Font(name="Calibri", size=14, bold=True, color=text_color)
            title_font = Font(name="Calibri", size=20, bold=True, color=primary_color)
            body_font = Font(name="Calibri", size=11, color=text_color)
            link_font = Font(name="Calibri", size=11, color="2563EB", underline="single")
            
            header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
            alt_fill = PatternFill(start_color=secondary_color, end_color=secondary_color, fill_type="solid")
            accent_fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type="solid")
            
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            
            # Create Overview Sheet
            overview_ws = wb.create_sheet("📋 Project Overview")
            self._create_overview_sheet(overview_ws, project_details, user_name, title_font, header_font, body_font, center_alignment, left_alignment)
            
            # Create Implementation Guide Sheet
            guide_ws = wb.create_sheet("📖 Implementation Guide")
            self._create_implementation_sheet(guide_ws, project_details, header_font, subheader_font, body_font, left_alignment)
            
            # Create Components Sheet
            components_ws = wb.create_sheet("🔧 Components")
            self._create_components_sheet(components_ws, project_details, header_font, body_font, header_fill, alt_fill, center_alignment, left_alignment, thin_border)
            
            # Create Frameworks & Tools Sheet
            frameworks_ws = wb.create_sheet("🛠️ Tools & Frameworks")
            self._create_frameworks_sheet(frameworks_ws, project_details, header_font, body_font, header_fill, alt_fill, center_alignment, left_alignment, thin_border)
            
            # Create Resources Sheet
            resources_ws = wb.create_sheet("📚 Learning Resources")
            self._create_resources_sheet(resources_ws, project_details, header_font, subheader_font, body_font, link_font, header_fill, alt_fill, center_alignment, left_alignment, thin_border)
            
            # Create Action Plan Sheet
            action_ws = wb.create_sheet("🎯 Action Plan")
            self._create_action_plan_sheet(action_ws, project_details, header_font, body_font, accent_fill, left_alignment, thin_border)
            
            # Set Overview as active sheet
            wb.active = overview_ws
            
            # Save to BytesIO buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            print(f"Error generating Excel file: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _create_overview_sheet(self, ws, project_details, user_name, title_font, header_font, body_font, center_alignment, left_alignment):
        """Create the project overview sheet"""
        
        # Title section
        ws.merge_cells("A1:F3")
        ws["A1"] = "🚀 ProjectCraft AI - Project Guide"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_alignment
        
        # Project title
        ws.merge_cells("A5:F6")
        ws["A5"] = project_details.title
        ws["A5"].font = Font(name="Calibri", size=18, bold=True, color="1E40AF")
        ws["A5"].alignment = center_alignment
        
        # Project info table
        current_time = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        
        info_data = [
            ["Project Information", "Details"],
            ["Personal Guide for:", user_name],
            ["Difficulty Level:", project_details.difficulty_level],
            ["Estimated Timeline:", project_details.estimated_time],
            ["Generated on:", current_time]
        ]
        
        start_row = 8
        for i, (label, value) in enumerate(info_data):
            ws[f"B{start_row + i}"] = label
            ws[f"D{start_row + i}"] = value
            
            if i == 0:  # Header row
                ws[f"B{start_row + i}"].font = header_font
                ws[f"D{start_row + i}"].font = header_font
                ws[f"B{start_row + i}"].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                ws[f"D{start_row + i}"].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                ws[f"B{start_row + i}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
                ws[f"D{start_row + i}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            else:
                ws[f"B{start_row + i}"].font = Font(name="Calibri", size=11, bold=True)
                ws[f"D{start_row + i}"].font = body_font
        
        # Project description
        ws[f"B{start_row + 6}"] = "Project Description:"
        ws[f"B{start_row + 6}"].font = Font(name="Calibri", size=12, bold=True, color="1E40AF")
        
        ws.merge_cells(f"B{start_row + 7}:F{start_row + 10}")
        ws[f"B{start_row + 7}"] = project_details.short_description
        ws[f"B{start_row + 7}"].font = body_font
        ws[f"B{start_row + 7}"].alignment = left_alignment
        
        # Quick stats
        ws[f"B{start_row + 12}"] = "Quick Project Stats:"
        ws[f"B{start_row + 12}"].font = Font(name="Calibri", size=12, bold=True, color="1E40AF")
        
        components_count = len(project_details.components) if project_details.components else 0
        frameworks_count = len(project_details.frameworks) if project_details.frameworks else 0
        youtube_count = len([link for link in project_details.youtube_links if link.strip()]) if project_details.youtube_links else 0
        github_count = len([repo for repo in project_details.github_repos if repo.strip()]) if project_details.github_repos else 0
        
        stats_data = [
            ["📦 Components Required:", str(components_count)],
            ["🛠️ Tools & Frameworks:", str(frameworks_count)],
            ["🎥 Video Tutorials:", str(youtube_count)],
            ["💻 GitHub Repositories:", str(github_count)]
        ]
        
        for i, (label, value) in enumerate(stats_data):
            ws[f"B{start_row + 14 + i}"] = label
            ws[f"D{start_row + 14 + i}"] = value
            ws[f"B{start_row + 14 + i}"].font = body_font
            ws[f"D{start_row + 14 + i}"].font = Font(name="Calibri", size=11, bold=True, color="10B981")
        
        # Set column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 5
        ws.column_dimensions["F"].width = 20
        
        # Set row heights
        for row in range(1, 25):
            ws.row_dimensions[row].height = 20
        
        ws.row_dimensions[7].height = 80  # Description row
    
    def _create_implementation_sheet(self, ws, project_details, header_font, subheader_font, body_font, left_alignment):
        """Create the implementation guide sheet"""
        
        # Title
        ws.merge_cells("A1:E2")
        ws["A1"] = "📖 Detailed Implementation Guide"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        # Process detailed description
        if project_details.detailed_description:
            lines = project_details.detailed_description.split('\n')
            current_row = 4
            
            for line in lines:
                line = line.strip()
                if not line:
                    current_row += 1
                    continue
                    
                if line.startswith('## '):
                    # Sub-heading
                    clean_line = line.replace('## ', '').strip()
                    ws.merge_cells(f"A{current_row}:E{current_row}")
                    ws[f"A{current_row}"] = clean_line
                    ws[f"A{current_row}"].font = subheader_font
                    ws[f"A{current_row}"].fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                    current_row += 1
                elif line.startswith('# '):
                    # Main heading
                    clean_line = line.replace('# ', '').strip()
                    ws.merge_cells(f"A{current_row}:E{current_row}")
                    ws[f"A{current_row}"] = clean_line
                    ws[f"A{current_row}"].font = Font(name="Calibri", size=14, bold=True, color="1E40AF")
                    ws[f"A{current_row}"].fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
                    current_row += 1
                elif line.startswith(('- ', '• ', '* ')):
                    # Bullet point
                    clean_line = line[2:].strip()
                    ws[f"B{current_row}"] = f"• {clean_line}"
                    ws[f"B{current_row}"].font = body_font
                    ws[f"B{current_row}"].alignment = left_alignment
                    current_row += 1
                elif line.startswith(tuple(f"{i}." for i in range(1, 10))):
                    # Numbered list
                    ws[f"B{current_row}"] = line
                    ws[f"B{current_row}"].font = body_font
                    ws[f"B{current_row}"].alignment = left_alignment
                    current_row += 1
                else:
                    # Normal paragraph
                    if len(line) > 10:
                        ws.merge_cells(f"A{current_row}:E{current_row}")
                        ws[f"A{current_row}"] = line
                        ws[f"A{current_row}"].font = body_font
                        ws[f"A{current_row}"].alignment = left_alignment
                        current_row += 1
        else:
            # Fallback content
            ws.merge_cells("A4:E6")
            ws["A4"] = "Detailed implementation steps will be provided based on your specific project requirements and chosen technologies."
            ws["A4"].font = body_font
            ws["A4"].alignment = left_alignment
        
        # Set column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
    
    def _create_components_sheet(self, ws, project_details, header_font, body_font, header_fill, alt_fill, center_alignment, left_alignment, thin_border):
        """Create the components sheet"""
        
        # Title
        ws.merge_cells("A1:D2")
        ws["A1"] = "🔧 Required Components & Specifications"
        ws["A1"].font = header_font
        ws["A1"].alignment = center_alignment
        
        # Headers
        headers = ["Component Name", "Purpose", "Specifications", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(4, col, header)
            cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Data rows
        if project_details.components and len(project_details.components) > 0:
            for row, comp in enumerate(project_details.components, 5):
                name = comp.get('name', 'Component')
                purpose = comp.get('purpose', 'Project component')
                specs = comp.get('specs', 'As per requirements')
                notes = "Research suppliers for best prices"
                
                data = [name, purpose, specs, notes]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row, col, value)
                    cell.font = body_font
                    cell.alignment = left_alignment
                    cell.border = thin_border
                    
                    if row % 2 == 0:
                        cell.fill = alt_fill
        else:
            # Fallback message
            ws.merge_cells("A5:D7")
            ws["A5"] = ("Components will be determined based on your specific project requirements. "
                       "Research the necessary hardware/software components for your implementation.")
            ws["A5"].font = body_font
            ws["A5"].alignment = left_alignment
        
        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 25
    
    def _create_frameworks_sheet(self, ws, project_details, header_font, body_font, header_fill, alt_fill, center_alignment, left_alignment, thin_border):
        """Create the frameworks and tools sheet"""
        
        # Title
        ws.merge_cells("A1:C2")
        ws["A1"] = "🛠️ Recommended Tools & Frameworks"
        ws["A1"].font = header_font
        ws["A1"].alignment = center_alignment
        
        # Headers
        headers = ["Tool/Framework", "Category", "Purpose"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(4, col, header)
            cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Data rows
        if project_details.frameworks and len(project_details.frameworks) > 0:
            for row, framework in enumerate(project_details.frameworks, 5):
                # Categorize frameworks
                framework_lower = framework.lower()
                if any(term in framework_lower for term in ['python', 'javascript', 'react', 'node']):
                    category = "Programming"
                elif any(term in framework_lower for term in ['arduino', 'iot', 'sensor']):
                    category = "Hardware/IoT"
                elif any(term in framework_lower for term in ['database', 'sql', 'mongodb']):
                    category = "Database"
                elif any(term in framework_lower for term in ['design', 'cad', 'modeling']):
                    category = "Design/Modeling"
                else:
                    category = "Development Tool"
                
                purpose = f"Essential for {category.lower()} aspects of the project"
                
                data = [framework, category, purpose]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row, col, value)
                    cell.font = body_font
                    cell.alignment = left_alignment
                    cell.border = thin_border
                    
                    if row % 2 == 0:
                        cell.fill = alt_fill
        else:
            # Fallback message
            ws.merge_cells("A5:C7")
            ws["A5"] = ("Recommended tools and frameworks will depend on your specific project requirements. "
                       "Consider popular development environments and libraries in your chosen field.")
            ws["A5"].font = body_font
            ws["A5"].alignment = left_alignment
        
        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 40
    
    def _create_resources_sheet(self, ws, project_details, header_font, subheader_font, body_font, link_font, header_fill, alt_fill, center_alignment, left_alignment, thin_border):
        """Create the learning resources sheet"""
        
        # Title
        ws.merge_cells("A1:D2")
        ws["A1"] = "📚 Learning Resources & References"
        ws["A1"].font = header_font
        ws["A1"].alignment = center_alignment
        
        current_row = 4
        
        # YouTube Tutorials Section
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "🎥 Video Tutorials"
        ws[f"A{current_row}"].font = subheader_font
        ws[f"A{current_row}"].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        current_row += 2
        
        # YouTube headers
        youtube_headers = ["#", "Tutorial Title/Search Term", "Link", "Type"]
        for col, header in enumerate(youtube_headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        if project_details.youtube_links and any(link.strip() for link in project_details.youtube_links if link):
            video_count = 0
            for link in project_details.youtube_links:
                if link and link.strip():
                    video_count += 1
                    if link.startswith('http'):
                        title = f"Direct Tutorial Link {video_count}"
                        link_text = link.strip()
                        link_type = "Direct Link"
                    else:
                        title = link.strip()
                        link_text = f"Search YouTube for: {link.strip()}"
                        link_type = "Search Term"
                    
                    data = [str(video_count), title, link_text, link_type]
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(current_row, col, value)
                        if col == 3 and link.startswith('http'):  # Link column
                            cell.font = link_font
                        else:
                            cell.font = body_font
                        cell.alignment = left_alignment
                        cell.border = thin_border
                        
                        if current_row % 2 == 0:
                            cell.fill = alt_fill
                    
                    current_row += 1
        else:
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = "Search YouTube for tutorials related to your project components and implementation steps."
            ws[f"A{current_row}"].font = body_font
            ws[f"A{current_row}"].alignment = left_alignment
            current_row += 1
        
        current_row += 2
        
        # GitHub Repositories Section  
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "💻 Code Repositories"
        ws[f"A{current_row}"].font = subheader_font
        ws[f"A{current_row}"].fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
        current_row += 2
        
        # GitHub headers
        github_headers = ["#", "Repository Name/Search Term", "Link", "Type"]
        for col, header in enumerate(github_headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid")
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        if project_details.github_repos and any(repo.strip() for repo in project_details.github_repos if repo):
            repo_count = 0
            for repo in project_details.github_repos:
                if repo and repo.strip():
                    repo_count += 1
                    if repo.startswith('http'):
                        repo_name = repo.split('/')[-1] if '/' in repo else f"Repository {repo_count}"
                        link_text = repo.strip()
                        repo_type = "Direct Link"
                    else:
                        repo_name = repo.strip()
                        link_text = f"Search GitHub for: {repo.strip()}"
                        repo_type = "Search Term"
                    
                    data = [str(repo_count), repo_name, link_text, repo_type]
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(current_row, col, value)
                        if col == 3 and repo.startswith('http'):  # Link column
                            cell.font = link_font
                        else:
                            cell.font = body_font
                        cell.alignment = left_alignment
                        cell.border = thin_border
                        
                        if current_row % 2 == 0:
                            cell.fill = alt_fill
                    
                    current_row += 1
        else:
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = "Search GitHub for open-source projects and code examples related to your project."
            ws[f"A{current_row}"].font = body_font
            ws[f"A{current_row}"].alignment = left_alignment
        
        # Set column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 15
    
    def _create_action_plan_sheet(self, ws, project_details, header_font, body_font, accent_fill, left_alignment, thin_border):
        """Create the action plan sheet"""
        
        # Title
        ws.merge_cells("A1:D2")
        ws["A1"] = "🎯 Project Action Plan & Next Steps"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        # Action steps
        action_steps = [
            {
                "phase": "Phase 1: Planning & Preparation",
                "tasks": [
                    "Gather all required components from the components list",
                    "Set up development environment and install necessary tools",
                    "Review all tutorial videos and documentation",
                    "Create project folder structure and documentation"
                ]
            },
            {
                "phase": "Phase 2: Learning & Research",
                "tasks": [
                    "Watch key tutorial videos to understand concepts",
                    "Study relevant GitHub repositories for code examples",
                    "Practice with smaller components before full integration",
                    "Join relevant online communities and forums"
                ]
            },
            {
                "phase": "Phase 3: Implementation",
                "tasks": [
                    "Start with basic functionality implementation",
                    "Test each component individually before integration",
                    "Follow step-by-step implementation guide",
                    "Document your progress and code changes"
                ]
            },
            {
                "phase": "Phase 4: Testing & Refinement",
                "tasks": [
                    "Perform comprehensive testing of all features",
                    "Debug issues and optimize performance",
                    "Add additional features and improvements",
                    "Prepare project documentation and presentation"
                ]
            }
        ]
        
        current_row = 4
        
        for phase_data in action_steps:
            # Phase header
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = phase_data["phase"]
            ws[f"A{current_row}"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            ws[f"A{current_row}"].fill = accent_fill
            ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"A{current_row}"].border = thin_border
            current_row += 1
            
            # Phase tasks
            for i, task in enumerate(phase_data["tasks"], 1):
                ws[f"A{current_row}"] = f"{i}."
                ws[f"B{current_row}"] = task
                ws[f"C{current_row}"] = "⬜ To Do"  # Checkbox
                ws[f"D{current_row}"] = "Notes"
                
                ws[f"A{current_row}"].font = Font(name="Calibri", size=11, bold=True)
                ws[f"B{current_row}"].font = body_font
                ws[f"C{current_row}"].font = Font(name="Calibri", size=11, color="10B981")
                ws[f"D{current_row}"].font = Font(name="Calibri", size=10, italic=True, color="6B7280")
                
                ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="top")
                ws[f"B{current_row}"].alignment = left_alignment
                ws[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="top")
                ws[f"D{current_row}"].alignment = left_alignment
                
                for col in ["A", "B", "C", "D"]:
                    ws[f"{col}{current_row}"].border = thin_border
                
                current_row += 1
            
            current_row += 1  # Space between phases
        
        # Tips section
        current_row += 1
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "💡 Tips for Success"
        ws[f"A{current_row}"].font = Font(name="Calibri", size=14, bold=True, color="1E40AF")
        ws[f"A{current_row}"].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        tips = [
            "Start with the basics and build incrementally",
            "Don't hesitate to ask for help in online communities",
            "Document your progress and learnings for future reference",
            "Test each component before integrating into the main system",
            "Have fun and be creative with your implementation!"
        ]
        
        for tip in tips:
            ws[f"A{current_row}"] = "💡"
            ws[f"B{current_row}"] = tip
            ws[f"A{current_row}"].font = Font(name="Calibri", size=12)
            ws[f"B{current_row}"].font = body_font
            ws[f"B{current_row}"].alignment = left_alignment
            current_row += 1
        
        # Set column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

    async def get_component_info(self, components: List[Dict]) -> List[str]:
        """Get component purchase links and information"""
        try:
            if not components:
                return []
                
            component_links = []
            for comp in components[:5]:  # Limit to 5 components
                component_name = comp.get("name", "")
                if component_name:
                    tavily_tool = self.tool_map.get("component_info_search")
                    if tavily_tool:
                        try:
                            # Search for where to buy this component
                            search_query = f"where to buy {component_name} electronics component online"
                            result = await asyncio.to_thread(
                                tavily_tool.invoke, {"query": search_query}
                            )
                            if result and len(result) > 20:  # Ensure we got useful results
                                component_links.append(f"**{component_name}**: {result[:200]}...")
                        except Exception as e:
                            component_links.append(f"**{component_name}**: Search online stores like Amazon, Adafruit, SparkFun, or local electronics suppliers.")
            
            return component_links
        except Exception as e:
            st.error(f"Error fetching component information: {e}")
        return []

def create_streamlit_app():
    """Create the enhanced Streamlit interface"""
    st.set_page_config(
        page_title="ProjectCraft AI - Your Intelligent Project Guide",
        page_icon="🚀",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Add custom CSS from theme
    add_custom_css()
    
    # Create animated title from theme
    create_animated_title()
    
    # Initialize session state
    if "conversation_history" not in st.session_state:
        st.session_state.conversation_history = []
    if "project_details" not in st.session_state:
        st.session_state.project_details = None
    if "current_stage" not in st.session_state:
        st.session_state.current_stage = "idea_input"
    if "assistant" not in st.session_state:
        st.session_state.assistant = ProjectGuideAssistant()
    if "user_name" not in st.session_state:
        st.session_state.user_name = ""
    if "selected_field" not in st.session_state:
        st.session_state.selected_field = None
    if "selected_subdomain" not in st.session_state:
        st.session_state.selected_subdomain = None
    if "selected_project" not in st.session_state:
        st.session_state.selected_project = None
    if "project_type" not in st.session_state:
        st.session_state.project_type = None
    if "complexity_level" not in st.session_state:
        st.session_state.complexity_level = None
    if "trending_projects" not in st.session_state:
        st.session_state.trending_projects = []
    if "refinement_questions" not in st.session_state:
        st.session_state.refinement_questions = []
    if "user_responses" not in st.session_state:
        st.session_state.user_responses = {}

    # Create progress indicator from theme - only show if user has started
    if st.session_state.conversation_history or st.session_state.current_stage != "idea_input":
        create_progress_indicator(st.session_state.current_stage)
    
    # Create interactive assistant from theme
    create_interactive_assistant(st.session_state.current_stage)

    # Enhanced Sidebar using theme function
    with st.sidebar:
        create_sidebar_stages(st.session_state.current_stage)
        
        st.markdown("---")
        
        # User profile section
        if not st.session_state.user_name:
            st.session_state.user_name = st.text_input("👋 What's your name?", placeholder="Enter your name")
        else:
            st.markdown(f"""
            <div style="background: linear-gradient(45deg, #667eea, #764ba2); 
                        border-radius: 10px; padding: 1rem; color: white; text-align: center;">
                <h4>👋 Welcome, {st.session_state.user_name}!</h4>
                <p style="margin: 0; opacity: 0.9;">Let's build something amazing together!</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        if st.button("🔄 Start Fresh Journey", use_container_width=True):
            # Clear all session state for a fresh start
            keys_to_clear = [
                "conversation_history", "project_details", "current_stage", "assistant",
                "selected_field", "selected_subdomain", "selected_project", 
                "project_type", "complexity_level", "trending_projects",
                "refinement_questions", "user_responses", "component_info"
            ]
            for key in keys_to_clear:
                if key in st.session_state:
                    del st.session_state[key]
            
            # Reset to initial stage
            st.session_state.current_stage = "idea_input"
            st.success("🎉 Starting fresh! Ready for your next amazing project!")
            st.rerun()

    # Main content area with enhanced styling
    if st.session_state.current_stage == "idea_input":
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #667eea; font-size: 2.5rem; margin-bottom: 0.5rem;">
                💡 What's Your Dream Project?
            </h2>
            <p style="font-size: 1.2rem; color: #6c757d; margin-bottom: 2rem;">
                Every great innovation starts with a simple idea. Share yours and let's make it extraordinary! ✨
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Engineering Fields Structure based on Pakistan's complete list
        engineering_fields = {
            "🏗️ Civil & Infrastructure": {
                "color": "#8B4513",
                "desc": "Building the foundations of society",
                "subfields": [
                    "Civil Engineering",
                    "Structural Engineering", 
                    "Environmental Engineering",
                    "Transportation Engineering",
                    "Water Resources Engineering",
                    "Geotechnical Engineering",
                    "Urban & Infrastructure Engineering",
                    "Architectural Engineering",
                    "Geological Engineering"
                ]
            },
            "⚙️ Mechanical & Manufacturing": {
                "color": "#FF6B35",
                "desc": "Designing and building mechanical systems",
                "subfields": [
                    "Mechanical Engineering",
                    "Manufacturing Engineering",
                    "Automotive Engineering",
                    "Aerospace Engineering",
                    "Aeronautical Engineering",
                    "Industrial Engineering",
                    "Mechatronics Engineering",
                    "Robotics & Automation Engineering",
                    "Marine Engineering"
                ]
            },
            "⚡ Electrical & Electronics": {
                "color": "#FFD700",
                "desc": "Powering the modern world",
                "subfields": [
                    "Electrical Engineering",
                    "Electronics Engineering",
                    "Telecommunication Engineering",
                    "Avionics Engineering",
                    "Power Systems Engineering",
                    "Control Systems Engineering",
                    "Instrumentation Engineering",
                    "Energy & Power Engineering"
                ]
            },
            "💻 Computing & Software": {
                "color": "#4169E1",
                "desc": "Creating digital solutions",
                "subfields": [
                    "Computer Engineering",
                    "Software Engineering",
                    "Information & Communication Technology (ICT)",
                    "Systems Engineering",
                    "Artificial Intelligence Engineering",
                    "Data Science Engineering",
                    "Cybersecurity Engineering",
                    "Network Engineering"
                ]
            },
            "⚗️ Chemical & Materials": {
                "color": "#32CD32",
                "desc": "Transforming materials and processes",
                "subfields": [
                    "Chemical Engineering",
                    "Petroleum Engineering",
                    "Polymer Engineering",
                    "Materials Engineering",
                    "Nanotechnology Engineering",
                    "Process Engineering",
                    "Food Engineering",
                    "Textile Engineering"
                ]
            },
            "🌱 Biological & Agricultural": {
                "color": "#228B22",
                "desc": "Engineering for life sciences",
                "subfields": [
                    "Biomedical Engineering",
                    "Agricultural Engineering",
                    "Bioengineering",
                    "Biotechnology Engineering",
                    "Food Engineering",
                    "Environmental Engineering",
                    "Biosystems Engineering"
                ]
            },
            "🛡️ Specialized & Defense": {
                "color": "#800080",
                "desc": "Advanced and specialized fields",
                "subfields": [
                    "Nuclear Engineering",
                    "Military Engineering",
                    "Defense Production Engineering",
                    "Mining Engineering",
                    "Metallurgical Engineering",
                    "Safety Engineering",
                    "Quality Engineering"
                ]
            },
            "🔬 Emerging Technologies": {
                "color": "#FF1493",
                "desc": "Cutting-edge engineering disciplines",
                "subfields": [
                    "Renewable Energy Engineering",
                    "Artificial Intelligence Engineering",
                    "Quantum Engineering",
                    "Space Technology Engineering",
                    "Biomedical Device Engineering",
                    "Smart Systems Engineering",
                    "Sustainable Engineering"
                ]
            }
        }

        # Interactive Field Selection UI
        st.markdown("""
        <h3 style="text-align: center; color: #495057; margin: 2rem 0 1rem 0;">
            🎓 Choose Your Engineering Field
        </h3>
        """, unsafe_allow_html=True)

        # Step 1: Select Main Field
        if not st.session_state.selected_field:
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        border-radius: 15px; padding: 1.5rem; margin: 1rem 0; 
                        text-align: center; color: white;">
                <h4 style="margin: 0;">📚 Step 1: Select Your Engineering Domain</h4>
                <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">Choose the main field that interests you most</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Create field selection grid
            cols = st.columns(2)
            for i, (field_name, field_info) in enumerate(engineering_fields.items()):
                with cols[i % 2]:
                    if st.button(
                        f"{field_name}\n{field_info['desc']}", 
                        key=f"field_{i}",
                        use_container_width=True,
                        help=f"Explore {len(field_info['subfields'])} specializations in this field"
                    ):
                        st.session_state.selected_field = field_name
                        # Clear old data when field changes
                        st.session_state.trending_projects = []
                        st.session_state.selected_project = None
                        st.session_state.project_details = None
                        st.session_state.user_responses = {}
                        st.rerun()

        # Step 2: Select Subdomain
        elif st.session_state.selected_field and not st.session_state.selected_subdomain:
            selected_field_info = engineering_fields[st.session_state.selected_field]
            
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, {selected_field_info['color']} 0%, {selected_field_info['color']}CC 100%); 
                        border-radius: 15px; padding: 1.5rem; margin: 1rem 0; 
                        text-align: center; color: white;">
                <h4 style="margin: 0;">🎯 Step 2: Choose Your Specialization</h4>
                <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">Selected: {st.session_state.selected_field}</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Back button
            if st.button("← Back to Fields", key="back_to_fields"):
                st.session_state.selected_field = None
                st.session_state.selected_subdomain = None
                # Clear old data when going back
                st.session_state.trending_projects = []
                st.session_state.selected_project = None
                st.session_state.project_details = None
                st.session_state.user_responses = {}
                st.rerun()
            
            st.markdown("### Available Specializations:")
            
            # Create subdomain selection
            subdomain = st.selectbox(
                "Choose your specific engineering specialization:",
                [""] + selected_field_info['subfields'],
                key="subdomain_select",
                help="Select the specific engineering discipline you want to focus on"
            )
            
            if subdomain:
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    if st.button("✅ Confirm Selection", type="primary", use_container_width=True):
                        st.session_state.selected_subdomain = subdomain
                        st.session_state.user_input = f"I want to build a project in {subdomain}"
                        # Clear old trending projects when subdomain changes
                        st.session_state.trending_projects = []
                        st.session_state.selected_project = None
                        st.session_state.project_details = None
                        st.session_state.user_responses = {}
                        st.balloons()
                        st.rerun()

        # Step 3: Confirmation and proceed
        elif st.session_state.selected_field and st.session_state.selected_subdomain:
            selected_field_info = engineering_fields[st.session_state.selected_field]
            
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); 
                        border-radius: 15px; padding: 2rem; margin: 1rem 0; 
                        text-align: center; color: white; box-shadow: 0 10px 30px rgba(0,0,0,0.1);">
                <div style="font-size: 3rem; margin-bottom: 0.5rem;">🎉</div>
                <h3 style="margin: 0;">Perfect Choice!</h3>
                <h4 style="margin: 0.5rem 0;">{st.session_state.selected_subdomain}</h4>
                <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">in {st.session_state.selected_field}</p>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🔄 Choose Different Field", use_container_width=True):
                    st.session_state.selected_field = None
                    st.session_state.selected_subdomain = None
                    # Clear all related data when changing field
                    st.session_state.trending_projects = []
                    st.session_state.selected_project = None
                    st.session_state.project_details = None
                    st.session_state.user_responses = {}
                    st.rerun()
            
            with col2:
                if st.button("🚀 Show Trending Projects", type="primary", use_container_width=True):
                    st.session_state.current_stage = "project_suggestions"
                    st.rerun()

        # Alternative: Direct Text Input Option
        if not st.session_state.selected_field:
            st.markdown("<br><hr><br>", unsafe_allow_html=True)
            
            st.markdown("""
            <div style="background: white; border-radius: 15px; padding: 1.5rem; 
                        box-shadow: 0 8px 25px rgba(0,0,0,0.1); margin: 2rem 0;">
                <h4 style="color: #495057; margin-bottom: 1rem;">💭 Or describe your project idea directly!</h4>
                <p style="color: #6c757d; font-size: 0.9rem; margin-bottom: 1rem;">
                    Have a specific project in mind? Skip the categories and tell me directly what you want to build!
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            user_input = st.text_area(
                "Describe your project idea:",
                placeholder="🌟 Example: I'm interested in making something that can help monitor my plants, or I want to learn about AI and build something cool with it...",
                height=120,
                key="user_input",
                help="Share your interests, what you'd like to learn, or any project ideas you have in mind!",
                label_visibility="collapsed"
            )
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("🚀 Start Chatting", type="primary", use_container_width=True) and user_input:
                    st.session_state.conversation_history.append(f"User: {user_input}")
                    st.session_state.current_stage = "refinement"
                    st.rerun()

    elif st.session_state.current_stage == "project_suggestions":
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #667eea; font-size: 2.5rem; margin-bottom: 0.5rem;">
                🔥 Trending Projects in Your Field
            </h2>
            <p style="font-size: 1.2rem; color: #6c757d; margin-bottom: 2rem;">
                Here are some popular and industry-relevant projects in {field_name}! ✨
            </p>
        </div>
        """.format(field_name=st.session_state.selected_subdomain or st.session_state.selected_field), unsafe_allow_html=True)
        
        # Generate trending projects if not already done
        if not st.session_state.trending_projects:
            with st.spinner("🔍 Finding trending projects in your field..."):
                field_for_projects = st.session_state.selected_subdomain or st.session_state.selected_field
                trending_projects = asyncio.run(
                    st.session_state.assistant.generate_trending_projects(field_for_projects)
                )
                st.session_state.trending_projects = trending_projects
        
        # Display trending projects
        if st.session_state.trending_projects:
            st.markdown("### 🎯 Select a Project That Interests You:")
            
            # Create project cards
            cols = st.columns(2)
            for i, project in enumerate(st.session_state.trending_projects):
                with cols[i % 2]:
                    # Determine card color based on difficulty
                    if project['difficulty'] == 'Beginner':
                        card_color = "#4CAF50"
                    elif project['difficulty'] == 'Intermediate':
                        card_color = "#FF9800"
                    else:
                        card_color = "#F44336"
                    
                    st.markdown(f"""
                    <div style="background: white; border-radius: 15px; padding: 1.5rem; margin: 1rem 0;
                                box-shadow: 0 8px 25px rgba(0,0,0,0.1); border-left: 5px solid {card_color};">
                        <h4 style="color: #495057; margin-bottom: 0.5rem;">{project['title']}</h4>
                        <p style="color: #6c757d; font-size: 0.9rem; margin-bottom: 1rem;">{project['description']}</p>
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 1rem;">
                            <span style="background: {card_color}; color: white; padding: 0.25rem 0.75rem; 
                                        border-radius: 15px; font-size: 0.8rem; font-weight: 600;">
                                {project['difficulty']}
                            </span>
                            <span style="color: #667eea; font-size: 0.8rem; font-weight: 500;">
                                {project['category']}
                            </span>
                        </div>
                        <div style="margin-bottom: 1rem;">
                            <strong style="color: #495057; font-size: 0.9rem;">Key Technologies:</strong>
                            <div style="margin-top: 0.5rem;">
                    """, unsafe_allow_html=True)
                    
                    # Display technologies as badges
                    tech_badges = ""
                    for tech in project['key_technologies']:
                        tech_badges += f'<span style="background: #e9ecef; color: #495057; padding: 0.2rem 0.5rem; margin: 0.2rem; border-radius: 10px; font-size: 0.75rem; display: inline-block;">{tech}</span>'
                    
                    st.markdown(f"""
                            {tech_badges}
                            </div>
                        </div>
                        <p style="color: #28a745; font-size: 0.8rem; font-style: italic; margin-bottom: 1rem;">
                            💡 {project['why_trending']}
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    if st.button(f"Select: {project['title']}", key=f"select_project_{i}", use_container_width=True):
                        st.session_state.selected_project = project
                        st.session_state.current_stage = "project_type_selection"
                        st.rerun()
        
        # Option for custom project idea
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    border-radius: 15px; padding: 1.5rem; margin: 2rem 0; 
                    text-align: center; color: white;">
            <h4 style="margin: 0;">💭 Have Your Own Project Idea?</h4>
            <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">No problem! You can describe your own project idea instead.</p>
        </div>
        """, unsafe_allow_html=True)
        
        custom_project = st.text_area(
            "Describe your custom project idea:",
            placeholder="e.g., I want to build a smart irrigation system that uses AI to predict watering needs...",
            height=100,
            help="Describe what you want to build, what problem it solves, or what you want to learn!"
        )
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back to Field Selection", use_container_width=True):
                st.session_state.selected_field = None
                st.session_state.selected_subdomain = None
                st.session_state.current_stage = "idea_input"
                # Clear all project-related data when going back
                st.session_state.trending_projects = []
                st.session_state.selected_project = None
                st.session_state.project_details = None
                st.session_state.user_responses = {}
                st.rerun()
        
        with col2:
            if st.button("🚀 Use Custom Idea", type="primary", use_container_width=True) and custom_project:
                # Create custom project object
                st.session_state.selected_project = {
                    "title": "Custom Project",
                    "description": custom_project,
                    "difficulty": "Custom",
                    "category": "Custom Project",
                    "key_technologies": [],
                    "why_trending": "Your unique idea!"
                }
                st.session_state.current_stage = "project_type_selection"
                st.rerun()

    elif st.session_state.current_stage == "project_type_selection":
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #667eea; font-size: 2.5rem; margin-bottom: 0.5rem;">
                🎯 Tell Me About Your Project
            </h2>
            <p style="font-size: 1.2rem; color: #6c757d; margin-bottom: 2rem;">
                Let's understand what kind of project you want to build! 📋
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Show selected project
        if st.session_state.selected_project:
            project = st.session_state.selected_project
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); 
                        border-radius: 15px; padding: 1.5rem; margin: 1rem 0; 
                        text-align: center; color: white;">
                <h3 style="margin: 0;">Selected Project: {project['title']}</h3>
                <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">{project['description']}</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Project type selection
        st.markdown("### 📚 What type of project is this?")
        project_type = st.radio(
            "Select the category that best fits your project:",
            ["🎓 Semester Project", "🏆 Final Year Project (FYP)", "🛠️ Hobby Project", "💼 Industry Project"],
            help="This helps me understand the scope and complexity level you're aiming for"
        )
        
        # Complexity level selection
        st.markdown("### 🎯 What complexity level do you prefer?")
        complexity_level = st.radio(
            "Choose based on your current skill level and learning goals:",
            ["🌱 Beginner - I'm new to this field", 
             "🔧 Intermediate - I have some experience", 
             "🚀 Advanced - I want a challenging project"],
            help="This helps me ask the right questions and suggest appropriate components"
        )
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back to Projects", use_container_width=True):
                st.session_state.current_stage = "project_suggestions"
                # Clear project-specific data when going back
                st.session_state.project_type = None
                st.session_state.complexity_level = None
                st.session_state.user_responses = {}
                st.rerun()
        
        with col2:
            if st.button("✅ Start Project Refinement", type="primary", use_container_width=True):
                st.session_state.project_type = project_type
                st.session_state.complexity_level = complexity_level
                st.session_state.current_stage = "refinement"
                st.rerun()

    elif st.session_state.current_stage == "refinement":
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #667eea; font-size: 2.5rem; margin-bottom: 0.5rem;">
                🔍 Let's Refine Your Project Details
            </h2>
            <p style="font-size: 1.2rem; color: #6c757d; margin-bottom: 2rem;">
                I'll ask specific questions to help you build the perfect project! 🎯
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Show project context
        if st.session_state.selected_project:
            project = st.session_state.selected_project
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        border-radius: 15px; padding: 1.5rem; margin: 1rem 0; 
                        text-align: center; color: white;">
                <h3 style="margin: 0;">🎯 {project['title']}</h3>
                <p style="margin: 0.5rem 0;">📋 {st.session_state.project_type} • 🎯 {st.session_state.complexity_level}</p>
                <p style="margin: 0.5rem 0 0 0; opacity: 0.9; font-size: 0.9rem;">{project['description']}</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Initialize refinement questions if not started
        if not st.session_state.refinement_questions:
            with st.spinner("🤔 Preparing the first question..."):
                question = asyncio.run(
                    st.session_state.assistant.ask_refinement_question(
                        st.session_state.selected_project['title'],
                        st.session_state.selected_subdomain or st.session_state.selected_field,
                        st.session_state.project_type,
                        st.session_state.complexity_level,
                        st.session_state.user_responses
                    )
                )
                st.session_state.refinement_questions.append(question)
        
        # Display conversation
        st.markdown("### 💬 Project Refinement Discussion")
        
        for i, question in enumerate(st.session_state.refinement_questions):
            # Display question
            st.markdown(f"""
            <div style="background: #e3f2fd; border-radius: 15px; padding: 1rem; margin: 1rem 0; border-left: 4px solid #2196F3;">
                <strong>🤖 Assistant:</strong> {question}
            </div>
            """, unsafe_allow_html=True)
            
            # Display user response if exists
            if f"question_{i}" in st.session_state.user_responses:
                response = st.session_state.user_responses[f"question_{i}"]
                st.markdown(f"""
                <div style="background: #e8f5e8; border-radius: 15px; padding: 1rem; margin: 1rem 0; border-left: 4px solid #4CAF50;">
                    <strong>� You:</strong> {response}
                </div>
                """, unsafe_allow_html=True)
        
        # Current question input
        current_question_index = len([k for k in st.session_state.user_responses.keys() if k.startswith("question_")])
        
        if current_question_index < len(st.session_state.refinement_questions):
            st.markdown("### ✏️ Your Response:")
            user_response = st.text_area(
                f"Answer for question {current_question_index + 1}:",
                placeholder="Share your thoughts, preferences, or requirements...",
                height=100,
                key=f"response_input_{current_question_index}",
                help="Be as specific as possible - this helps me generate a better project guide!"
            )
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("📝 Submit Response", type="primary", use_container_width=True) and user_response:
                    # Save the response
                    st.session_state.user_responses[f"question_{current_question_index}"] = user_response
                    
                    # Generate next question if we haven't asked enough
                    if len(st.session_state.refinement_questions) < 4:  # Ask up to 4 questions
                        with st.spinner("🤔 Thinking of the next question..."):
                            next_question = asyncio.run(
                                st.session_state.assistant.ask_refinement_question(
                                    st.session_state.selected_project['title'],
                                    st.session_state.selected_subdomain or st.session_state.selected_field,
                                    st.session_state.project_type,
                                    st.session_state.complexity_level,
                                    st.session_state.user_responses
                                )
                            )
                            st.session_state.refinement_questions.append(next_question)
                    
                    st.rerun()
        
        # Show completion option when enough questions are answered
        if len(st.session_state.user_responses) >= 3:  # After 3 questions minimum
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown("""
            <div style="background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); 
                        border-radius: 15px; padding: 1.5rem; margin: 1rem 0; 
                        text-align: center; color: white;">
                <h4 style="margin: 0;">🎉 Great! I have enough details to create your project guide!</h4>
                <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">Ready to see your personalized project blueprint?</p>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🔄 Ask More Questions", use_container_width=True):
                    # Generate one more question
                    with st.spinner("🤔 Preparing another question..."):
                        next_question = asyncio.run(
                            st.session_state.assistant.ask_refinement_question(
                                st.session_state.selected_project['title'],
                                st.session_state.selected_subdomain or st.session_state.selected_field,
                                st.session_state.project_type,
                                st.session_state.complexity_level,
                                st.session_state.user_responses
                            )
                        )
                        st.session_state.refinement_questions.append(next_question)
                    st.rerun()
            
            with col2:
                if st.button("🚀 Generate Project Guide", type="primary", use_container_width=True):
                    st.session_state.current_stage = "details"
                    st.rerun()
        
        # Back button
        if st.button("← Back to Project Selection"):
            st.session_state.current_stage = "project_type_selection"
            # Clear refinement data when going back
            st.session_state.refinement_questions = []
            st.session_state.user_responses = {}
            st.rerun()
        
        # Get assistant response with loading animation
        if st.session_state.conversation_history:
            latest_user_input = st.session_state.conversation_history[-1]
            if latest_user_input.startswith("User:") and len([m for m in st.session_state.conversation_history if m.startswith("Assistant:")]) < len([m for m in st.session_state.conversation_history if m.startswith("User:")]):
                with st.spinner("🤔 Thinking about your question and preparing a helpful response..."):
                    try:
                        response = asyncio.run(
                            st.session_state.assistant.refine_project_idea(
                                latest_user_input[5:],
                                st.session_state.conversation_history
                            )
                        )
                        st.session_state.conversation_history.append(f"Assistant: {response}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Oops! I had trouble processing that. Let's try again: {e}")
        
        # Enhanced input section
        st.markdown("""
        <div style="background: white; border-radius: 15px; padding: 1.5rem; 
                    box-shadow: 0 8px 25px rgba(0,0,0,0.1); margin: 2rem 0;">
            <h4 style="color: #495057; margin-bottom: 1rem;">💭 Continue Our Chat</h4>
            <p style="color: #6c757d; font-size: 0.9rem; margin-bottom: 1rem;">
                Ask me anything! Share more about your interests, ask about technologies, or explore different project ideas.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        follow_up = st.text_input(
            "", 
            placeholder="Ask me anything about the project, share your doubts, or explore new ideas...",
            key="follow_up",
            help="Feel free to ask questions, share concerns, or explore different directions!"
        )
        
        col1, col2 = st.columns(2)
        


    elif st.session_state.current_stage == "details":
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #667eea; font-size: 2.5rem; margin-bottom: 0.5rem;">
                📋 Crafting Your Project Blueprint
            </h2>
            <p style="font-size: 1.2rem; color: #6c757d; margin-bottom: 2rem;">
                Time to transform your idea into a detailed, actionable plan! 🎯
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.project_details is None:
            # Extract project title from conversation
            project_title = "Custom Project"
            for msg in st.session_state.conversation_history:
                if "project" in msg.lower():
                    project_title = msg.split(":")[-1].strip()[:50]
                    break
            
            # Enhanced loading with progress simulation
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            loading_steps = [
                "🔍 Analyzing your requirements...",
                "🧠 Generating project structure...",
                "⚙️ Selecting components...",
                "📚 Finding frameworks...",
                "🎯 Finalizing details..."
            ]
            
            try:
                for i, step in enumerate(loading_steps):
                    status_text.text(step)
                    progress_bar.progress((i + 1) / len(loading_steps))
                    
                if not st.session_state.project_details:  # Only generate if not already generated
                    st.session_state.project_details = asyncio.run(
                        st.session_state.assistant.generate_project_details(project_title)
                    )
                
                progress_bar.progress(1.0)
                status_text.success("✅ Your project blueprint is ready!")
                st.balloons()
                
            except Exception as e:
                st.error(f"Oops! Encountered an issue: {e}")
                return
        
        # Display project details with enhanced styling
        details = st.session_state.project_details
        if details:
            # Project header with gradient background
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        border-radius: 15px; padding: 2rem; margin: 2rem 0; 
                        text-align: center; color: white; box-shadow: 0 10px 30px rgba(0,0,0,0.1);">
                <h1 style="margin: 0; font-size: 2.5rem; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">
                    🎯 {details.title}
                </h1>
                <p style="margin: 1rem 0 0 0; font-size: 1.2rem; opacity: 0.9;">
                    {details.short_description}
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            # Project metadata cards
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"""
                <div class="project-card" style="text-align: center;">
                    <div style="font-size: 2rem; margin-bottom: 0.5rem;">📊</div>
                    <h4 style="color: #495057; margin: 0;">Difficulty</h4>
                    <p style="color: #667eea; font-weight: 600; font-size: 1.1rem; margin: 0.5rem 0 0 0;">
                        {details.difficulty_level}
                    </p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="project-card" style="text-align: center;">
                    <div style="font-size: 2rem; margin-bottom: 0.5rem;">⏱️</div>
                    <h4 style="color: #495057; margin: 0;">Timeline</h4>
                    <p style="color: #667eea; font-weight: 600; font-size: 1.1rem; margin: 0.5rem 0 0 0;">
                        {details.estimated_time}
                    </p>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="project-card" style="text-align: center;">
                    <div style="font-size: 2rem; margin-bottom: 0.5rem;">🔧</div>
                    <h4 style="color: #495057; margin: 0;">Components</h4>
                    <p style="color: #667eea; font-weight: 600; font-size: 1.1rem; margin: 0.5rem 0 0 0;">
                        {len(details.components)} Items
                    </p>
                </div>
                """, unsafe_allow_html=True)
            
            # Detailed guide section with better formatting
            st.markdown("""
            <div class="project-card">
                <h3 style="color: #495057; margin-bottom: 1rem;">📖 Step-by-Step Implementation Guide</h3>
            </div>
            """, unsafe_allow_html=True)
            
            # Enhanced detailed description formatting
            if details.detailed_description:
                # Clean and format the description
                description = details.detailed_description.strip()
                
                # Split into sections and format properly
                sections = description.split('\n')
                formatted_content = []
                
                for line in sections:
                    line = line.strip()
                    if not line:
                        continue
                    
                    if line.startswith('## '):
                        # Sub-heading
                        heading = line.replace('## ', '').strip()
                        formatted_content.append(f"""
                        <div style="background: linear-gradient(45deg, #667eea, #764ba2); 
                                    border-radius: 8px; padding: 0.75rem; margin: 1rem 0; color: white;">
                            <h4 style="margin: 0; font-size: 1.1rem;">📋 {heading}</h4>
                        </div>
                        """)
                    elif line.startswith('# '):
                        # Main heading  
                        heading = line.replace('# ', '').strip()
                        formatted_content.append(f"""
                        <div style="background: linear-gradient(45deg, #4CAF50, #45a049); 
                                    border-radius: 10px; padding: 1rem; margin: 1.5rem 0; color: white; text-align: center;">
                            <h3 style="margin: 0; font-size: 1.3rem;">🎯 {heading}</h3>
                        </div>
                        """)
                    elif line.startswith(('- ', '• ', '*')):
                        # Bullet points
                        bullet_text = line[2:].strip()
                        formatted_content.append(f"""
                        <div style="background: #f8f9fa; border-left: 4px solid #667eea; 
                                    padding: 0.75rem; margin: 0.5rem 0;">
                            <p style="margin: 0; color: #495057;">• {bullet_text}</p>
                        </div>
                        """)
                    elif line.startswith(tuple(f"{i}." for i in range(1, 10))):
                        # Numbered lists
                        formatted_content.append(f"""
                        <div style="background: #fff3e0; border-left: 4px solid #FF9800; 
                                    padding: 0.75rem; margin: 0.5rem 0;">
                            <p style="margin: 0; color: #495057; font-weight: 500;">{line}</p>
                        </div>
                        """)
                    else:
                        # Regular paragraph
                        if len(line) > 10:  # Only substantial content
                            formatted_content.append(f"""
                            <div style="background: white; border-radius: 8px; padding: 1rem; 
                                        margin: 0.75rem 0; box-shadow: 0 2px 8px rgba(0,0,0,0.05);">
                                <p style="margin: 0; color: #495057; line-height: 1.6;">{line}</p>
                            </div>
                            """)
                
                # Display all formatted content
                for content in formatted_content:
                    st.markdown(content, unsafe_allow_html=True)
            else:
                # Fallback if no detailed description
                st.markdown("""
                <div style="background: white; border-radius: 15px; padding: 2rem; 
                            box-shadow: 0 8px 25px rgba(0,0,0,0.1); margin: 2rem 0; text-align: center;">
                    <h4 style="color: #495057;">📋 Project Guide Coming Soon</h4>
                    <p style="color: #6c757d;">We're working on creating a detailed guide for this project.</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Components section with enhanced cards
            if details.components:
                st.markdown("""
                <div class="project-card">
                    <h3 style="color: #495057; margin-bottom: 1rem;">🔧 Required Components</h3>
                </div>
                """, unsafe_allow_html=True)
                
                # Display components in a grid layout
                if len(details.components) <= 3:
                    cols = st.columns(len(details.components))
                else:
                    cols = st.columns(3)
                
                for i, comp in enumerate(details.components):
                    with cols[i % len(cols)]:
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #667eea, #764ba2); 
                                    border-radius: 15px; padding: 1.5rem; margin: 1rem 0; 
                                    color: white; box-shadow: 0 6px 20px rgba(102,126,234,0.3);
                                    min-height: 200px;">
                            <div style="text-align: center; margin-bottom: 1rem;">
                                <div style="font-size: 2.5rem; margin-bottom: 0.5rem;">�</div>
                                <h4 style="margin: 0; font-size: 1.1rem;">{comp.get('name', 'Component')}</h4>
                            </div>
                            <div style="background: rgba(255,255,255,0.1); border-radius: 8px; padding: 1rem;">
                                <p style="margin: 0 0 0.75rem 0; font-size: 0.9rem; font-weight: 500;">
                                    <strong>Purpose:</strong>
                                </p>
                                <p style="margin: 0 0 1rem 0; font-size: 0.85rem; opacity: 0.9;">
                                    {comp.get('purpose', 'Essential component for the project')}
                                </p>
                                <p style="margin: 0 0 0.5rem 0; font-size: 0.9rem; font-weight: 500;">
                                    <strong>Specs:</strong>
                                </p>
                                <p style="margin: 0; font-size: 0.85rem; opacity: 0.9;">
                                    {comp.get('specs', 'Standard specifications available')}
                                </p>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
            
            # Frameworks section
            if details.frameworks:
                st.markdown("""
                <div class="project-card">
                    <h3 style="color: #495057; margin-bottom: 1rem;">🛠️ Recommended Tools & Frameworks</h3>
                </div>
                """, unsafe_allow_html=True)
                
                framework_cols = st.columns(min(len(details.frameworks), 4))
                for i, framework in enumerate(details.frameworks):
                    with framework_cols[i % len(framework_cols)]:
                        st.markdown(f"""
                        <div style="background: linear-gradient(45deg, #4ECDC4, #44A08D); 
                                    border-radius: 10px; padding: 1rem; text-align: center; 
                                    color: white; margin: 0.5rem 0;">
                            <strong>{framework}</strong>
                        </div>
                        """, unsafe_allow_html=True)
            
            # Action button
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("🔗 Find Resources & Tutorials", type="primary", use_container_width=True):
                    st.session_state.current_stage = "resources"
                    st.success("🎉 Let's gather all the resources you need!")
                    st.rerun()

    elif st.session_state.current_stage == "resources":
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #667eea; font-size: 2.5rem; margin-bottom: 0.5rem;">
                🔗 Your Learning Resources Hub
            </h2>
            <p style="font-size: 1.2rem; color: #6c757d; margin-bottom: 2rem;">
                Everything you need to bring your project to life - tutorials, code, and components! 📚
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        details = st.session_state.project_details
        if details:
            # Resources grid layout
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("""
                <div class="project-card">
                    <h3 style="color: #495057; margin-bottom: 1rem;">
                        📺 Video Tutorials
                    </h3>
                </div>
                """, unsafe_allow_html=True)
                
                if details.youtube_links:
                    for i, link in enumerate(details.youtube_links):
                        if link:  # Only display if link exists
                            if link.startswith('http'):
                                # Actual YouTube link
                                st.markdown(f"""
                                <div style="background: linear-gradient(135deg, #FF6B6B, #FF5252); 
                                            border-radius: 12px; padding: 1.2rem; margin: 0.8rem 0; 
                                            color: white; box-shadow: 0 4px 15px rgba(255,107,107,0.3);
                                            transition: transform 0.2s;">
                                    <a href="{link}" target="_blank" style="color: white; text-decoration: none;">
                                        <div style="display: flex; align-items: center; gap: 10px;">
                                            <div style="font-size: 1.5rem;">🎬</div>
                                            <div>
                                                <strong>Tutorial Video {i+1}</strong><br>
                                                <small style="opacity: 0.9;">Click to watch comprehensive tutorial</small>
                                            </div>
                                        </div>
                                    </a>
                                </div>
                                """, unsafe_allow_html=True)
                            else:
                                # YouTube search URL
                                st.markdown(f"""
                                <div style="background: linear-gradient(135deg, #FFA726, #FF7043); 
                                            border-radius: 12px; padding: 1.2rem; margin: 0.8rem 0; 
                                            color: white; box-shadow: 0 4px 15px rgba(255,167,38,0.3);">
                                    <a href="{link}" target="_blank" style="color: white; text-decoration: none;">
                                        <div style="display: flex; align-items: center; gap: 10px;">
                                            <div style="font-size: 1.5rem;">🔍</div>
                                            <div>
                                                <strong>Search Tutorial {i+1}</strong><br>
                                                <small style="opacity: 0.9;">Click to search YouTube</small>
                                            </div>
                                        </div>
                                    </a>
                                </div>
                                """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div style="background: #e3f2fd; border: 2px dashed #2196F3; 
                                border-radius: 10px; padding: 1.5rem; text-align: center; margin: 1rem 0;">
                        <div style="font-size: 2rem; margin-bottom: 0.5rem;">🔍</div>
                        <p style="color: #1976D2; font-weight: 500; margin: 0;">
                            Search YouTube for comprehensive tutorials on your project topic!
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
            
            with col2:
                st.markdown("""
                <div class="project-card">
                    <h3 style="color: #495057; margin-bottom: 1rem;">
                        💻 GitHub Repositories
                    </h3>
                </div>
                """, unsafe_allow_html=True)
                
                if details.github_repos:
                    for i, repo in enumerate(details.github_repos):
                        if repo:  # Only display if repo exists
                            if repo.startswith('http'):
                                # Actual GitHub repository link
                                repo_name = repo.split('/')[-1] if '/' in repo else f"Repository {i+1}"
                                st.markdown(f"""
                                <div style="background: linear-gradient(135deg, #4ECDC4, #44A08D); 
                                            border-radius: 12px; padding: 1.2rem; margin: 0.8rem 0; 
                                            color: white; box-shadow: 0 4px 15px rgba(78,205,196,0.3);
                                            transition: transform 0.2s;">
                                    <a href="{repo}" target="_blank" style="color: white; text-decoration: none;">
                                        <div style="display: flex; align-items: center; gap: 10px;">
                                            <div style="font-size: 1.5rem;">⭐</div>
                                            <div>
                                                <strong>{repo_name}</strong><br>
                                                <small style="opacity: 0.9;">Source code and documentation</small>
                                            </div>
                                        </div>
                                    </a>
                                </div>
                                """, unsafe_allow_html=True)
                            else:
                                # GitHub search URL
                                st.markdown(f"""
                                <div style="background: linear-gradient(135deg, #66BB6A, #4CAF50); 
                                            border-radius: 12px; padding: 1.2rem; margin: 0.8rem 0; 
                                            color: white; box-shadow: 0 4px 15px rgba(102,187,106,0.3);">
                                    <a href="{repo}" target="_blank" style="color: white; text-decoration: none;">
                                        <div style="display: flex; align-items: center; gap: 10px;">
                                            <div style="font-size: 1.5rem;">🔍</div>
                                            <div>
                                                <strong>Search Repository {i+1}</strong><br>
                                                <small style="opacity: 0.9;">Click to search GitHub</small>
                                            </div>
                                        </div>
                                    </a>
                                </div>
                                """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div style="background: #e8f5e8; border: 2px dashed #4CAF50; 
                                border-radius: 10px; padding: 1.5rem; text-align: center; margin: 1rem 0;">
                        <div style="font-size: 2rem; margin-bottom: 0.5rem;">🔍</div>
                        <p style="color: #2E7D32; font-weight: 500; margin: 0;">
                            Search GitHub for open-source implementations of similar projects!
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
            
            # Component shopping section
            st.markdown("""
            <div class="project-card">
                <h3 style="color: #495057; margin-bottom: 1rem;">
                    🛒 Component Shopping Guide
                </h3>
            </div>
            """, unsafe_allow_html=True)
            
            if details.components:
                st.markdown("""
                <div style="background: linear-gradient(45deg, #FECA57, #FF9FF3); 
                            border-radius: 10px; padding: 1.5rem; color: white; text-align: center; margin-bottom: 1rem;">
                    <h4 style="margin: 0;">💡 Where to Buy Components</h4>
                    <p style="margin: 0.5rem 0 0 0;">
                        Find these components at electronics stores like Amazon, Adafruit, SparkFun, or local suppliers
                    </p>
                </div>
                """, unsafe_allow_html=True)
                
                # Display component info from get_component_info
                component_info = getattr(details, 'component_info', [])
                if hasattr(st.session_state, 'component_info') and st.session_state.component_info:
                    for info in st.session_state.component_info:
                        st.markdown(f"""
                        <div style="background: white; border: 2px solid #667eea; 
                                    border-radius: 10px; padding: 1rem; margin: 0.5rem 0;">
                            {info}
                        </div>
                        """, unsafe_allow_html=True)
                
                # Always show component list
                component_cols = st.columns(min(len(details.components), 3))
                for i, comp in enumerate(details.components):
                    with component_cols[i % len(component_cols)]:
                        st.markdown(f"""
                        <div style="background: white; border: 2px solid #667eea; 
                                    border-radius: 10px; padding: 1rem; margin: 0.5rem 0; text-align: center;">
                            <h5 style="color: #667eea; margin: 0 0 0.5rem 0;">
                                {comp.get('name', 'Component')}
                            </h5>
                            <p style="color: #6c757d; font-size: 0.9rem; margin: 0;">
                                {comp.get('purpose', 'Essential component')}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
            else:
                st.info("ℹ️ This project mainly requires software tools and frameworks.")
            
            # Learning resources summary
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        border-radius: 15px; padding: 2rem; margin: 2rem 0; 
                        text-align: center; color: white; box-shadow: 0 10px 30px rgba(0,0,0,0.1);">
                <h3 style="margin: 0 0 1rem 0;">🎓 Additional Learning Resources</h3>
                <div style="display: flex; justify-content: space-around; flex-wrap: wrap;">
                    <div style="margin: 0.5rem;">
                        <div style="font-size: 2rem;">📖</div>
                        <strong>Documentation</strong>
                    </div>
                    <div style="margin: 0.5rem;">
                        <div style="font-size: 2rem;">🎯</div>
                        <strong>Tutorials</strong>
                    </div>
                    <div style="margin: 0.5rem;">
                        <div style="font-size: 2rem;">💬</div>
                        <strong>Community</strong>
                    </div>
                    <div style="margin: 0.5rem;">
                        <div style="font-size: 2rem;">🛠️</div>
                        <strong>Tools</strong>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Action button
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("📄 Create Final Project Guide", type="primary", use_container_width=True):
                    st.session_state.current_stage = "export"
                    st.success("🎉 Time to package everything into your personal guide!")
                    st.rerun()

    elif st.session_state.current_stage == "export":
        st.markdown("""
        <div style="text-align: center; margin: 2rem 0;">
            <h2 style="color: #667eea; font-size: 2.5rem; margin-bottom: 0.5rem;">
                📄 Your Complete Project Guide
            </h2>
            <p style="font-size: 1.2rem; color: #6c757d; margin-bottom: 2rem;">
                Congratulations! Your personalized project guide is ready to download! 🎉
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        details = st.session_state.project_details
        if details:
            # Success celebration
            st.balloons()
            
            # Project summary card
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); 
                        border-radius: 15px; padding: 2rem; margin: 2rem 0; 
                        text-align: center; color: white; box-shadow: 0 10px 30px rgba(0,0,0,0.1);">
                <div style="font-size: 4rem; margin-bottom: 1rem;">🎯</div>
                <h1 style="margin: 0; font-size: 2rem; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">
                    {details.title}
                </h1>
                <p style="margin: 1rem 0 0 0; font-size: 1.1rem; opacity: 0.9;">
                    Ready to build something amazing!
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            # Guide statistics
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown("""
                <div class="project-card" style="text-align: center;">
                    <div style="font-size: 2rem; color: #4CAF50;">📋</div>
                    <h4 style="color: #495057; margin: 0.5rem 0;">Complete Guide</h4>
                    <p style="color: #667eea; font-weight: 600; margin: 0;">Ready!</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="project-card" style="text-align: center;">
                    <div style="font-size: 2rem; color: #FF9800;">🔧</div>
                    <h4 style="color: #495057; margin: 0.5rem 0;">Components</h4>
                    <p style="color: #667eea; font-weight: 600; margin: 0;">{len(details.components)}</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="project-card" style="text-align: center;">
                    <div style="font-size: 2rem; color: #2196F3;">🔗</div>
                    <h4 style="color: #495057; margin: 0.5rem 0;">Resources</h4>
                    <p style="color: #667eea; font-weight: 600; margin: 0;">{len(details.youtube_links + details.github_repos)}</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                st.markdown(f"""
                <div class="project-card" style="text-align: center;">
                    <div style="font-size: 2rem; color: #9C27B0;">⏱️</div>
                    <h4 style="color: #495057; margin: 0.5rem 0;">Timeline</h4>
                    <p style="color: #667eea; font-weight: 600; margin: 0;">{details.estimated_time}</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Create downloadable summary with enhanced formatting
            user_name = st.session_state.get("user_name", "Builder")
            summary = f"""# 🚀 {details.title}
*Personal Project Guide for {user_name}*

---

## 📋 Project Overview

**Difficulty Level:** {details.difficulty_level}  
**Estimated Timeline:** {details.estimated_time}  
**Generated:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

## 💡 Project Description

{details.short_description}

## 📖 Detailed Implementation Guide

{details.detailed_description}

## 🔧 Required Components

"""
            for i, comp in enumerate(details.components, 1):
                summary += f"""{i}. **{comp.get('name', 'Component')}**
   - Purpose: {comp.get('purpose', 'N/A')}
   - Specifications: {comp.get('specs', 'N/A')}

"""
            
            summary += f"""## 🛠️ Recommended Tools & Frameworks

"""
            for framework in details.frameworks:
                summary += f"- {framework}\n"
            
            summary += f"""
## 📺 Learning Resources

### Video Tutorials
"""
            for i, link in enumerate(details.youtube_links, 1):
                summary += f"{i}. [Tutorial Video {i}]({link})\n"
            
            summary += f"""
### Code Repositories
"""
            for i, repo in enumerate(details.github_repos, 1):
                summary += f"{i}. [GitHub Repository {i}]({repo})\n"
            
            summary += f"""
---

## 🎯 Next Steps

1. **Gather Components** - Use the component list to purchase or gather required materials
2. **Study Resources** - Watch tutorials and explore code repositories
3. **Start Building** - Follow the step-by-step guide
4. **Join Community** - Connect with other builders and share your progress
5. **Iterate & Improve** - Make the project your own!

## 💡 Tips for Success

- Start with the basics and build incrementally
- Don't hesitate to ask for help in online communities
- Document your progress and learnings
- Test each component before integrating
- Have fun and be creative!

---

*Generated by ProjectCraft AI - Your Intelligent Project Guide*  
*Happy Building! 🚀*
"""
            
            # Download section
            st.markdown("""
            <div style="background: white; border-radius: 15px; padding: 2rem; 
                        box-shadow: 0 8px 25px rgba(0,0,0,0.1); margin: 2rem 0; text-align: center;">
                <h3 style="color: #495057; margin-bottom: 1rem;">📥 Download Your Guide</h3>
                <p style="color: #6c757d; margin-bottom: 2rem;">
                    Your complete project guide is ready! Choose your preferred format.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            # Download buttons in two columns
            col1, col2 = st.columns(2)
            
            with col1:
                # Excel Download
                with st.spinner("Generating Professional Excel File..."):
                    excel_data = st.session_state.assistant.generate_excel_guide(details, user_name)
                    if excel_data:
                        st.download_button(
                            label="📊 Download Excel Guide (Professional)",
                            data=excel_data,
                            file_name=f"{details.title.replace(' ', '_')}_ProjectGuide.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True,
                            help="Professional Excel format - Perfect for planning and tracking progress"
                        )
                    else:
                        st.error("Error generating Excel file. Please try again.")
            
            with col2:
                # Markdown Download (fallback)
                st.download_button(
                    label="📝 Download Markdown Guide",
                    data=summary,
                    file_name=f"{details.title.replace(' ', '_')}_ProjectGuide.md",
                    mime="text/markdown",
                    type="secondary",
                    use_container_width=True,
                    help="Markdown format - Easy to edit and version control"
                )
            
            # Celebration and next steps
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        border-radius: 15px; padding: 2rem; margin: 2rem 0; 
                        text-align: center; color: white; box-shadow: 0 10px 30px rgba(0,0,0,0.1);">
                <h3 style="margin: 0 0 1rem 0;">🎉 Congratulations!</h3>
                <p style="margin: 0; font-size: 1.1rem; opacity: 0.9;">
                    You now have everything you need to build an amazing project. 
                    Remember, every expert was once a beginner. Happy building! 🚀
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            # Action buttons
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🔄 Start Another Project", type="secondary", use_container_width=True):
                    # Clear all session state for a fresh start
                    keys_to_clear = [
                        "conversation_history", "project_details", "current_stage", "assistant",
                        "selected_field", "selected_subdomain", "selected_project", 
                        "project_type", "complexity_level", "trending_projects",
                        "refinement_questions", "user_responses", "component_info"
                    ]
                    for key in keys_to_clear:
                        if key in st.session_state:
                            del st.session_state[key]
                    
                    # Reset to initial stage
                    st.session_state.current_stage = "idea_input"
                    st.success("🎉 Ready for your next amazing project!")
                    st.rerun()
            
            with col2:
                if st.button("💬 Share Feedback", type="secondary", use_container_width=True):
                    st.info("💡 We'd love to hear about your project journey! Share your success stories!")

if __name__ == "__main__":
    create_streamlit_app()
